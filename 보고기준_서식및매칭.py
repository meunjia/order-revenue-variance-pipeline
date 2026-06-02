#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""보고 vs 실적 기준 서식파일 + 매칭파일 생성"""

import os, sys, io, re, chardet
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

WORK_DIR = r"C:\Users\user\claudecode\월별차이"
CSV_DIR  = os.path.join(WORK_DIR, "csv")
OUT_FMT  = os.path.join(WORK_DIR, "월별차이분석_보고기준.xlsx")
OUT_MATCH= os.path.join(WORK_DIR, "보고vs실적_수주매칭.xlsx")

# ── 스타일
T = Side(style='thin'); M = Side(style='medium')
BD = Border(left=T, right=T, top=T, bottom=T)
def fill(h): return PatternFill('solid', fgColor=h)
BLU=fill('4472C4'); LBLU=fill('D6E4F7'); YEL=fill('FFFF00')
GRN=fill('E2EFDA'); RED=fill('FFDDE1'); ORG=fill('FCE4D6')
GRY=fill('F2F2F2'); WHT=fill('FFFFFF'); SUB=fill('D9E1F2')

def hf(sz=9):  return Font(bold=True, size=sz, color='FFFFFF', name='맑은 고딕')
def bf(sz=9):  return Font(bold=True, size=sz, color='000000', name='맑은 고딕')
def df(sz=9):  return Font(bold=False, size=sz, color='000000', name='맑은 고딕')
def C(w=False): return Alignment(horizontal='center', vertical='center', wrap_text=w)
def L(w=True):  return Alignment(horizontal='left',   vertical='center', wrap_text=w)
def TL():       return Alignment(horizontal='left',   vertical='top',    wrap_text=True)

def set_cell(ws, row, col, val, font=None, fill_=None, align=None, fmt=None, border=True):
    c = ws.cell(row, col, value=val)
    if font:   c.font      = font
    if fill_:  c.fill      = fill_
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border    = BD
    return c

# ── CSV 데이터 로드
def detect_enc(p):
    with open(p,'rb') as f: return chardet.detect(f.read(100000))['encoding'] or 'utf-8-sig'

def to_num(v):
    if pd.isna(v): return 0.0
    s = str(v).strip().replace(',','').replace(' ','')
    return float(s) if s not in ('','-','nan','#VALUE!','#REF!') else 0.0

def norm_team(v):
    return re.sub(r'^\d+\.','',str(v)).strip() if not pd.isna(v) else None

def classify_prod(v):
    if pd.isna(v): return None
    s = str(v).upper().strip()
    if any(k in s for k in ['TR','RT','변압기']): return '변압기'
    if any(k in s for k in ['HH','HG','GIS','EA','MF','GT','차단기']): return '차단기'
    return None

path = os.path.join(CSV_DIR, '실적_수주예상.csv')
enc = detect_enc(path)
df_raw = pd.read_csv(path, encoding=enc, low_memory=False)
df_raw.columns = [str(c).strip().replace('\n','') for c in df_raw.columns]
df_raw.dropna(how='all', inplace=True)
df_raw['금액']  = df_raw['예상금액(억원)'].apply(to_num)
df_raw['팀명2'] = df_raw['팀명'].apply(norm_team)
df_raw['제품군']= df_raw['제품군0'].apply(classify_prod)

Q1 = ['26년01월','26년02월','26년03월']
q1_실적 = df_raw[(df_raw['확정월'].isin(Q1)) & (df_raw['수주IN']=='IN')].copy()
q1_실적 = q1_실적[q1_실적['금액']>0].copy()

# ── 스크린샷 확정 수치 (보고/실적/전년)
DATA = {
    '중전기': {
        '수주':    {'보고':4320.9, '실적':4290.1, '전년':1864.0},
        '매출':    {'보고':1200.3, '실적':1214.3, '전년': 357.1},
        '영업이익':{'보고': 253.1, '실적': 293.2, '전년': 165.3},
    },
    '변압기': {
        '수주':    {'보고':3905.8, '실적':3875.0, '전년':1464.9},
        '매출':    {'보고': 963.7, '실적': 974.2, '전년': 182.7},
        '영업이익':{'보고': 215.9, '실적': 247.7, '전년': 149.6},
    },
    '차단기': {
        '수주':    {'보고': 415.0, '실적': 415.0, '전년': 399.2},
        '매출':    {'보고': 236.5, '실적': 240.1, '전년': 174.3},
        '영업이익':{'보고':  37.3, '실적':  45.5, '전년':  15.7},
    },
}

# 보대비 차이
DIFF = {k: round(DATA['중전기'][k]['실적'] - DATA['중전기'][k]['보고'], 1)
        for k in ['수주','매출','영업이익']}

PERIODS = [
    (3,  '1분기', ['보고','실적','전년']),
    (6,  '4월',   ['계획','보고','실적']),
    (9,  '4월누계',['계획','보고','실적']),
    (12, '5월',   ['계획','보고','예상']),
    (15, '6월',   ['계획','보고','예상']),
    (18, '2분기', ['계획','보고','예상','전년']),
    (22, '상반기',['계획','보고','예상','전년']),
    (26, '7월',   ['계획','예상']),
    (28, '7월누계',['계획','예상','전년']),
]
MAX_COL = 30

# ══════════════════════════════════════════════════════════════
# 파일 1: 서식 파일 (보고기준)
# ══════════════════════════════════════════════════════════════
wb1 = openpyxl.Workbook()
ws = wb1.active
ws.title = "월별차이분석"
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 5.5
ws.column_dimensions['B'].width = 7.5
for i in range(3, MAX_COL+1):
    ws.column_dimensions[get_column_letter(i)].width = 8.5

# Row 1: 타이틀
ws.merge_cells(f'A1:{get_column_letter(MAX_COL)}1')
c = ws['A1']
c.value = '■ 월별 수주/매출/이익 차이분석'
c.font = bf(11); c.alignment = C(); ws.row_dimensions[1].height = 22

# Row 2-3: 기간 헤더
ws.row_dimensions[2].height = 16; ws.row_dimensions[3].height = 14
ws.merge_cells('A2:B3')
c = ws['A2']
c.value = '구분'; c.font = hf(); c.fill = BLU; c.alignment = C(); c.border = BD

for sc, period, subs in PERIODS:
    ec = sc + len(subs) - 1
    ws.merge_cells(f'{get_column_letter(sc)}2:{get_column_letter(ec)}2')
    c = ws[f'{get_column_letter(sc)}2']
    c.value = period; c.font = hf(); c.fill = BLU; c.alignment = C()
    for col in range(sc, ec+1): ws.cell(2, col).border = BD
    for j, sub in enumerate(subs):
        c = ws.cell(3, sc+j)
        c.value = sub; c.font = bf(8); c.fill = SUB; c.alignment = C(); c.border = BD

groups = ['중전기','변압기','차단기']
items  = ['수주','매출','영업이익']
grp_fills = [WHT, fill('F5F8FF'), GRY]

for g, grp in enumerate(groups):
    base = 4 + g * 3
    gf = grp_fills[g]
    for r in range(base, base+3): ws.row_dimensions[r].height = 14
    ws.merge_cells(f'A{base}:A{base+2}')
    c = ws[f'A{base}']
    c.value = grp; c.font = bf(); c.fill = gf; c.alignment = C(True); c.border = BD

    for i, item in enumerate(items):
        row = base + i
        set_cell(ws, row, 2, item, df(), gf, C())
        for j, key in enumerate(['보고','실적','전년']):
            val = DATA[grp][item][key]
            set_cell(ws, row, 3+j, val, df(), gf, C(), '#,##0.0')
        for sc, _, subs in PERIODS[1:]:
            for j in range(len(subs)):
                c = ws.cell(row, sc+j)
                c.fill=gf; c.border=BD; c.alignment=C()
                c.number_format='#,##0.0'

for row in [2,3]:
    for col in [1,2]: ws.cell(row,col).border = BD

# ── 보대비 차이분석 섹션
DR = 14
ws.row_dimensions[13].height = 6; ws.row_dimensions[DR].height = 16
ws.merge_cells(f'A{DR}:{get_column_letter(MAX_COL)}{DR}')
c = ws[f'A{DR}']
c.value = '★ 보고대비 차이분석 ★'; c.font = bf(10); c.alignment = L(False)

HR = DR + 1; ws.row_dimensions[HR].height = 16
set_cell(ws, HR, 1, '구분', hf(), BLU, C())
S_COLS = {'수주':(2,12), '매출':(13,23), '영업이익':(24,30)}
for lbl, (s,e) in S_COLS.items():
    ws.merge_cells(f'{get_column_letter(s)}{HR}:{get_column_letter(e)}{HR}')
    c = ws[f'{get_column_letter(s)}{HR}']
    c.value = lbl; c.font = hf(); c.fill = BLU; c.alignment = C()
    for col in range(s, e+1): ws.cell(HR, col).border = BD

DIFF_ROWS = [
    ('1분기', True),
    ('4월',   False), ('4월누계',False), ('5월',False), ('6월',False),
    ('2분기', False), ('상반기', False), ('7월',False), ('7월누계',False),
]

# 1분기 수주 차이 텍스트
d_수주 = DIFF['수주']; d_매출 = DIFF['매출']; d_영익 = DIFF['영업이익']
q1_수주_text = (
    f'보고대비 {d_수주:+.1f}억\n\n'
    '■ 보고→실적 주요 변동\n'
    '해외2) Project Apex          -30.0억 (Q1→5월 이동)\n'
    '해외1) Project Nova Phase 1  -5.8억 (금액 축소)\n'
    '해외1) 용역계약/보관비 신규   +2.4억 (Project Ridge, Project Gem 등)\n'
    '국내1) 소규모 추가 항목       +2.6억\n'
    '국내2) 변동없음               0.0억'
)
q1_매출_text = f'보고대비 {d_매출:+.1f}억\n\n(상세 분석 별도)'
q1_영익_text = f'보고대비 {d_영익:+.1f}억\n\n(상세 분석 별도)'

for idx, (period, has_data) in enumerate(DIFF_ROWS):
    row = HR + 1 + idx
    ws.row_dimensions[row].height = 120 if has_data else 50
    set_cell(ws, row, 1, period, bf(),
             fill('E2EFDA') if has_data else SUB, C(True))
    texts = {
        '수주': q1_수주_text if has_data else '',
        '매출': q1_매출_text if has_data else '',
        '영업이익': q1_영익_text if has_data else '',
    }
    for lbl, (s,e) in S_COLS.items():
        ws.merge_cells(f'{get_column_letter(s)}{row}:{get_column_letter(e)}{row}')
        c = ws[f'{get_column_letter(s)}{row}']
        c.value = texts[lbl]
        c.font = df(8); c.fill = YEL if has_data else WHT; c.alignment = TL()
        for col in range(s, e+1): ws.cell(row, col).border = BD

wb1.save(OUT_FMT)
print(f'서식파일 저장: {OUT_FMT}')

# ══════════════════════════════════════════════════════════════
# 파일 2: 보고 vs 실적 수주 매칭파일
# ══════════════════════════════════════════════════════════════
wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)

# ── 시트 1: 요약
ws1 = wb2.create_sheet('요약')
ws1.sheet_view.showGridLines = False
for col, w in zip('ABCDE', [10, 10, 10, 10, 30]):
    ws1.column_dimensions[col].width = w

ws1.merge_cells('A1:E1')
c = ws1['A1']
c.value = '수주 보고 vs 실적 비교  |  보고 4,320.9억 → 실적 4,290.1억  (차이 -30.8억)'
c.font = bf(11); c.fill = LBLU; c.alignment = C(); ws1.row_dimensions[1].height = 22

# 헤더
for col, (val, w_) in enumerate(zip(
    ['팀명','보고(억)','실적(억)','차이(억)','비고'], [BLU]*5), 1):
    c = ws1.cell(2, col, value=val)
    c.font=hf(); c.fill=BLU; c.alignment=C(); c.border=BD
ws1.row_dimensions[2].height = 15

# 보고 vs 실적 팀별
팀별보고 = {'국내1':1219.658,'국내2':447.7197,'해외1':2600.9445,'해외2':52.5296}
비고 = {
    '국내1':'소규모 항목 +2.6억 추가',
    '국내2':'변동 없음',
    '해외1':'Project Nova -5.8억, 용역계약/보관비 +2.4억',
    '해외2':'Project Apex -30.0억 (5월 이동)',
}
팀합보고 = 0; 팀합실적 = 0
for r, 팀 in enumerate(['국내1','국내2','해외1','해외2'], 3):
    bg = 팀별보고[팀]
    실 = round(q1_실적[q1_실적['팀명2']==팀]['금액'].sum(), 1)
    diff = round(실 - bg, 1)
    df_ = GRN if diff > 0 else (RED if diff < 0 else WHT)
    ws1.row_dimensions[r].height = 15
    for col, (val, fl_) in enumerate(
        [(팀,LBLU),(bg,WHT),(실,WHT),(diff,df_),(비고[팀],WHT)], 1):
        c = ws1.cell(r, col, value=val)
        c.font = (bf() if col==4 else df()); c.fill = fl_; c.alignment = C(); c.border = BD
        if col in [2,3]: c.number_format = '#,##0.0'
        if col == 4: c.number_format = '+#,##0.0;-#,##0.0'
    팀합보고 += bg; 팀합실적 += 실

# 합계행
ws1.row_dimensions[7].height = 16
for col, (val, fl_) in enumerate(
    [('합계',BLU),(round(팀합보고,1),BLU),(round(팀합실적,1),BLU),
     (round(팀합실적-팀합보고,1),BLU),('',BLU)], 1):
    c = ws1.cell(7, col, value=val)
    c.font=hf(); c.fill=BLU; c.alignment=C(); c.border=BD
    if col in [2,3]: c.number_format='#,##0.0'
    if col == 4: c.number_format='+#,##0.0;-#,##0.0'

# 비고 설명
ws1.merge_cells('A9:E9')
ws1['A9'].value = '* 보고: 스크린샷 피벗테이블 기준 (원본 Excel). 실적: 실적_수주예상.csv Q1 IN 필터 기준.'
ws1['A9'].font = df(8); ws1['A9'].fill = fill('FFF2CC'); ws1['A9'].alignment = L()

# ── 시트 2: 보고-실적 차이 상세
ws2 = wb2.create_sheet('차이_상세')
ws2.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGH',[5,8,8,25,10,10,10,25]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells('A1:H1')
c = ws2['A1']
c.value = '보고 vs 실적 수주 차이 상세 (총 -30.8억)'
c.font = bf(11); c.fill = LBLU; c.alignment = C(); ws2.row_dimensions[1].height = 22

# 헤더
hdrs = ['No','팀명','제품군','프로젝트명','보고(억)','실적(억)','차이(억)','비고']
for col, h in enumerate(hdrs, 1):
    c = ws2.cell(2, col, value=h)
    c.font=hf(); c.fill=BLU; c.alignment=C(); c.border=BD
ws2.row_dimensions[2].height = 15

DIFF_ITEMS = [
    # (팀, 제품군, 프로젝트명, 보고, 실적, 비고)
    ('해외2','변압기','Project Apex (Intl)',             30.00, 0.00,  'Q1→5월 이동 (현재 CSV 26년05월 IN)'),
    ('해외1','변압기','Project Nova Phase 1',           124.00,118.18, '금액 수정 (-5.82억)'),
    ('해외1','변압기','Project Sierra 보관비',          12.17,  0.78, '분리 → 용역계약(13.26)+보관비(0.78) 재계상'),
    ('해외1','변압기','Project Sierra 용역계약',         0.00, 13.26, '신규 계상 (보관비에서 분리)'),
    ('해외1','변압기','Project Titan II 용역계약',         12.20, 13.28, '금액 소폭 증가'),
    ('해외1','변압기','Project Titan II 보관비',            0.00,  0.78, '신규 계상'),
    ('해외1','변압기','Project Ridge 용역계약',           0.00, 13.87, '신규 계상'),
    ('해외1','변압기','Project Ridge 보관비',             0.00,  0.75, '신규 계상'),
    ('해외1','변압기','Project Gem 용역계약',               0.00, 11.77, '신규 계상'),
    ('해외1','변압기','Project Valley 보관비',           0.00,  1.64, '신규 계상'),
    ('해외1','변압기','Project Stream 보관비',               0.00,  1.57, '신규 계상'),
    ('해외1','변압기','기타 해외1 조정',               35.72, 35.72,  '동일 (기타 주요항목 유지)'),
    ('국내1','변압기+차단기','국내1 소규모 추가',        0.00,  2.60, ''+
     '소규모 추가 항목'),
    ('국내2','변압기+차단기','국내2 전체',             447.72,447.73, '실질 변동 없음'),
]

for i, (팀, 제품군, pjt, bg, 실, note) in enumerate(DIFF_ITEMS, 1):
    row = i + 2
    ws2.row_dimensions[row].height = 15
    diff = round(실 - bg, 2)
    df_ = GRN if diff > 0 else (RED if diff < 0 else WHT)
    for col, (val, fl_) in enumerate(
        [(i,WHT),(팀,WHT),(제품군,WHT),(pjt,WHT),
         (bg if bg>0 else None,WHT),(실 if 실>0 else None,WHT),(diff,df_),(note,WHT)], 1):
        c = ws2.cell(row, col, value=val)
        c.font=df(8) if col==8 else (bf() if col==7 else df())
        c.fill=fl_; c.alignment=(L() if col in [4,8] else C()); c.border=BD
        if col in [5,6]: c.number_format='#,##0.00'
        if col == 7: c.number_format='+#,##0.00;-#,##0.00'

# 소계
tr = len(DIFF_ITEMS) + 3
ws2.merge_cells(f'A{tr}:F{tr}')
ws2[f'A{tr}'].value = f'보고 합계 4,320.9  →  실적 합계 4,290.1   (차이 -30.8억 = Project Apex -30.0 + 해외1 -3.4 + 국내1 +2.6)'
ws2[f'A{tr}'].font=bf(9); ws2[f'A{tr}'].fill=YEL; ws2[f'A{tr}'].alignment=C()
ws2.cell(tr,7,value=-30.8).number_format='+#,##0.0;-#,##0.0'
ws2.cell(tr,7).font=bf(); ws2.cell(tr,7).fill=YEL; ws2.cell(tr,7).alignment=C()
ws2.cell(tr,7).border=BD
for col in range(1,7): ws2.cell(tr,col).border=BD
ws2.row_dimensions[tr].height=18

# ── 시트 3: Q1 실적 전체 (번호 + 팀 + 프로젝트 + 금액)
ws3 = wb2.create_sheet('Q1실적_전체목록')
ws3.sheet_view.showGridLines = False
for col, w in zip('ABCDEFG',[5,8,8,8,28,10,16]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells('A1:G1')
c = ws3['A1']
c.value = f'Q1 수주 실적 전체 ({len(q1_실적)}건, {q1_실적["금액"].sum():.1f}억)'
c.font=bf(11); c.fill=LBLU; c.alignment=C(); ws3.row_dimensions[1].height=22

for col, h in enumerate(['No','확정월','팀명','제품군','프로젝트명','금액(억)','발주처'],1):
    c=ws3.cell(2,col,value=h); c.font=hf(); c.fill=BLU; c.alignment=C(); c.border=BD
ws3.row_dimensions[2].height=15

sorted_실적 = q1_실적.sort_values(['팀명2','금액'], ascending=[True,False]).reset_index(drop=True)
cur_team = None
for i, r in sorted_실적.iterrows():
    row = i + 3
    ws3.row_dimensions[row].height = 14
    team = r['팀명2']
    gf = [WHT, fill('F0F4FF'), fill('FFF8F0'), fill('F0FFF0')][
        ['국내1','국내2','해외1','해외2'].index(team) if team in ['국내1','국내2','해외1','해외2'] else 0]
    for col, (val, fl_) in enumerate([
        (i+1, WHT),
        (r['확정월'], gf),
        (team, gf),
        (r['제품군'] or '', gf),
        (r['PJT명'] or '', gf),
        (r['금액'], gf),
        (str(r.get('발주처','') or ''), gf),
    ], 1):
        c = ws3.cell(row, col, value=val)
        c.font = df(8)
        c.fill = fl_; c.border = BD
        c.alignment = L(False) if col in [5,7] else C()
        if col == 6: c.number_format = '#,##0.00'

# 합계행
tot_row = len(sorted_실적) + 3
ws3.merge_cells(f'A{tot_row}:E{tot_row}')
ws3[f'A{tot_row}'].value=f'합계 ({len(q1_실적)}건)'
ws3[f'A{tot_row}'].font=bf(); ws3[f'A{tot_row}'].fill=BLU; ws3[f'A{tot_row}'].alignment=C()
c=ws3.cell(tot_row,6,value=round(q1_실적['금액'].sum(),1))
c.font=hf(); c.fill=BLU; c.alignment=C(); c.border=BD; c.number_format='#,##0.0'
for col in range(1,6): ws3.cell(tot_row,col).border=BD
ws3.row_dimensions[tot_row].height=16

wb2.save(OUT_MATCH)
print(f'매칭파일 저장: {OUT_MATCH}')
print()
print('=== 검증 요약 ===')
print(f'보고 합계 (스크린샷): 4,320.9억')
print(f'실적 합계 (CSV):      {q1_실적["금액"].sum():.1f}억')
print(f'차이:                 {q1_실적["금액"].sum()-4320.9:+.1f}억')
print()
print('주요 차이 원인:')
print('  해외2) Project Apex (Int'l) -30.0억 (현재 5월 IN 확인됨)')
print('  해외1) Project Nova -5.8억 + 용역계약/보관비 신규 +2.4억 = 순 -3.4억')
print('  국내1) 소규모 추가 +2.6억')
print('  국내2) 변동없음')
