#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl, sys, io, numpy as np, pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\user\claudecode\월별차이\차이원인분석.xlsx')

ws = wb.create_sheet('전체_수치흐름', 0)
ws.sheet_view.showGridLines = False

def F(hex): return PatternFill('solid', fgColor=hex)
def B():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

for i, w in enumerate([24,12,12,12,12,12,12,30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def C(r, c, v='', bold=False, fg='000000', size=10, bg=None, ha='center', wrap=False):
    cc = ws.cell(r, c, v)
    cc.font = Font(bold=bold, color=fg, size=size)
    cc.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cc.border = B()
    if bg: cc.fill = F(bg)
    return cc

def title_row(r, text, bg):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(r, 1, text)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = F(bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = B()
    ws.row_dimensions[r].height = 22
    return r + 1

def hdr(r, cols, bg='D9E1F2'):
    for ci, v in enumerate(cols, 1):
        C(r, ci, v, bold=True, bg=bg, size=9)
    ws.row_dimensions[r].height = 16
    return r + 1

def num_color(cc, v):
    if isinstance(v, (int, float)) and v != 0:
        if v > 0:  cc.font = Font(bold=True, color='0070C0', size=9)
        elif v < 0: cc.font = Font(bold=True, color='C00000', size=9)

PROD_BG = {'중전기': 'DEEAF1', '변압기': 'E2EFDA', '차단기': 'FFF2CC'}
CAT_BG  = {
    '신규': 'E2EFDA', '금액동일': 'FFFFFF', '금액변동': 'FFF2CC',
    '명칭변경추정': 'FCE4D6', '분할추정': 'DDEBF7', '제품군변경추정': 'EAD1DC',
    '소멸_대형(취소/이월추정)': 'FFDDE1', '소멸_소형(취소/명칭변경추정)': 'FFE9E9',
}
NOTE_MAP = {
    '신규':                        '계획에 없던 프로젝트 실적 발생',
    '금액동일':                     '계획=실적 완전 일치',
    '금액변동':                     '동일 프로젝트, 금액 변경',
    '명칭변경추정':                   '프로젝트명 변경 추정 (유사도 80%↑)',
    '분할추정':                     '계획 1건 → 실적 여러 건으로 분할',
    '제품군변경추정':                  '변압기↔차단기 제품군 변경',
    '소멸_대형(취소/이월추정)':         '10억↑ 프로젝트 실적 미발생 (취소/이월)',
    '소멸_소형(취소/명칭변경추정)':      '10억↓ 프로젝트 실적 미발생 (취소/명칭변경)',
}
CAT_ORDER = ['금액동일','금액변동','명칭변경추정','분할추정','제품군변경추정',
             '신규','소멸_대형(취소/이월추정)','소멸_소형(취소/명칭변경추정)']

row = 1

# ══════════════════════════════
# 섹션 1. 보고서 수치
# ══════════════════════════════
row = title_row(row, '① 보고서 수치 (제공된 원본)', '1F4E79')
row = hdr(row, ['구분','계획(억)','실적(억)','차이(억)','','','',''])
report = [
    ('중전기  수주', 2607.5, 4290.1),
    ('중전기  매출', 1140.5, 1214.3),
    ('변압기  수주', 2231.4, 3875.0),
    ('변압기  매출',  906.7,  974.2),
    ('차단기  수주',  376.2,  415.0),
    ('차단기  매출',  233.8,  240.1),
]
for nm, p, a in report:
    bg = PROD_BG.get(nm[:3], 'FFFFFF')
    diff = round(a - p, 1)
    for ci, v in enumerate([nm, p, a, diff, '', '', '', ''], 1):
        cc = C(row, ci, v, bg=bg, ha='left' if ci==1 else 'center', size=9)
        if ci == 4: num_color(cc, v)
    ws.row_dimensions[row].height = 14
    row += 1

row += 1  # 빈 줄

# ══════════════════════════════
# 섹션 2. CSV 계산 수치
# ══════════════════════════════
row = title_row(row, '② CSV 직접 집계 수치', '375623')
row = hdr(row, ['구분','계획(억)','실적(억)','차이(억)','','','',''])
csv_vals = [
    ('중전기  수주', 2232.6, 4320.9),
    ('중전기  매출', 1140.5, 1186.8),
    ('변압기  수주', 1981.3, 3905.8),
    ('변압기  매출',  906.7,  948.0),
    ('차단기  수주',  251.3,  415.1),
    ('차단기  매출',  233.8,  238.8),
]
for nm, p, a in csv_vals:
    bg = PROD_BG.get(nm[:3], 'FFFFFF')
    diff = round(a - p, 1)
    for ci, v in enumerate([nm, p, a, diff, '', '', '', ''], 1):
        cc = C(row, ci, v, bg=bg, ha='left' if ci==1 else 'center', size=9)
        if ci == 4: num_color(cc, v)
    ws.row_dimensions[row].height = 14
    row += 1

row += 1

# ══════════════════════════════
# 섹션 3. 보고서 vs CSV 갭
# ══════════════════════════════
row = title_row(row, '③ 보고서 vs CSV 갭 (미해결 데이터 차이)', 'C55A11')
row = hdr(row, ['구분','계획 갭(억)','실적 갭(억)','','비고','','',''])
gap_data = [
    ('중전기  수주',  374.9, -30.8, '수주계획 CSV 누락 / 실적 소폭 초과'),
    ('중전기  매출',    0.0, -27.5, '매출계획 일치 / 실적 소폭 미달'),
    ('변압기  수주',  250.1, -30.8, 'CSV 누락'),
    ('변압기  매출',    0.0, -26.2, ''),
    ('차단기  수주',  124.9,  -0.1, 'CSV 누락'),
    ('차단기  매출',    0.0,  -1.3, ''),
]
for nm, gp, ga, note in gap_data:
    bg = PROD_BG.get(nm[:3], 'FFFFFF')
    for ci, v in enumerate([nm, gp, ga, '', note, '', '', ''], 1):
        cc = C(row, ci, v, bg=bg, ha='left' if ci in (1,5) else 'center', size=9, wrap=(ci==5))
        if ci in (2, 3): num_color(cc, v)
    ws.row_dimensions[row].height = 14
    row += 1

row += 1

# ══════════════════════════════
# 섹션 4. 계획 vs 실적 차이 원인
# ══════════════════════════════
row = title_row(row, '④ 계획 vs 실적 차이 원인 분류 (CSV 기준)', '2E75B6')
row = hdr(row, ['원인분류','수주_건수','수주_차이(억)','└변압기','└차단기',
                '매출_건수','매출_차이(억)','설명'])

def read_df(ws2):
    hdrs = [ws2.cell(2, c).value for c in range(1, ws2.max_column+1)]
    rows = []
    for r2 in range(3, ws2.max_row+1):
        rv = [ws2.cell(r2, c).value for c in range(1, ws2.max_column+1)]
        if any(v is not None for v in rv):
            rows.append(dict(zip(hdrs, rv)))
    df = pd.DataFrame(rows)
    for col in ['계획금액','실적금액','차이(억)']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df

df_o = read_df(wb['수주_차이원인'])
df_s = read_df(wb['매출_차이원인'])

go = df_o.groupby('원인분류').agg(cnt=('차이(억)','count'), tot=('차이(억)','sum')).round(1)
go['tr'] = df_o[df_o['제품군']=='변압기'].groupby('원인분류')['차이(억)'].sum().round(1)
go['cb'] = df_o[df_o['제품군']=='차단기'].groupby('원인분류')['차이(억)'].sum().round(1)
go = go.fillna(0)

gs = df_s.groupby('원인분류').agg(cnt=('차이(억)','count'), tot=('차이(억)','sum')).round(1).fillna(0)

for cat in CAT_ORDER:
    if cat not in go.index and cat not in gs.index: continue
    oc  = int(go.loc[cat,'cnt']) if cat in go.index else 0
    ot  = go.loc[cat,'tot'] if cat in go.index else 0.0
    otr = go.loc[cat,'tr']  if cat in go.index else 0.0
    ocb = go.loc[cat,'cb']  if cat in go.index else 0.0
    sc  = int(gs.loc[cat,'cnt']) if cat in gs.index else 0
    st  = gs.loc[cat,'tot'] if cat in gs.index else 0.0
    note = NOTE_MAP.get(cat, '')
    bg = CAT_BG.get(cat, 'FFFFFF')

    for ci, v in enumerate([cat, oc, ot, otr, ocb, sc, st, note], 1):
        cc = C(row, ci, v, bold=(ci==1), bg=bg,
               ha='left' if ci in (1,8) else 'center', size=9, wrap=(ci==8))
        if ci in (3,4,5,7): num_color(cc, v)
    ws.row_dimensions[row].height = 14
    row += 1

# 합계행
for ci, v in enumerate(['합계',
    int(df_o['차이(억)'].count()), round(df_o['차이(억)'].sum(),1),
    round(df_o[df_o['제품군']=='변압기']['차이(억)'].sum(),1),
    round(df_o[df_o['제품군']=='차단기']['차이(억)'].sum(),1),
    int(df_s['차이(억)'].count()), round(df_s['차이(억)'].sum(),1), ''], 1):
    cc = C(row, ci, v, bold=True, bg='D9E1F2', ha='left' if ci==1 else 'center', size=9)
    if ci in (3,4,5,7): num_color(cc, v)
ws.row_dimensions[row].height = 16

ws.freeze_panes = 'A2'
wb.save(r'C:\Users\user\claudecode\월별차이\차이원인분석.xlsx')
print('저장 완료')
