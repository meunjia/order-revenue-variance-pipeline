#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""월별 수주/매출/이익 차이분석 보고서 양식 생성"""

import openpyxl, sys, io, pandas as pd, numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUT = r'C:\Users\user\claudecode\월별차이\월별차이분석_1분기.xlsx'
SRC = r'C:\Users\user\claudecode\월별차이\차이원인분석.xlsx'

# ── 스타일 헬퍼
def F(h): return PatternFill('solid', fgColor=h)
def Bd(thin=True):
    s = Side(style='thin' if thin else 'medium')
    return Border(left=s, right=s, top=s, bottom=s)
def Bd2():
    m = Side(style='medium'); t = Side(style='thin')
    return Border(left=m, right=m, top=t, bottom=t)

YELLOW = 'FFFF00'
HEADER = 'D9E1F2'
LIGHT  = 'F2F2F2'
WHITE  = 'FFFFFF'
NAVY   = '1F3864'

def cell(ws, r, c, v='', bold=False, fg='000000', size=9,
         bg=None, ha='center', va='center', wrap=False, border=True):
    cc = ws.cell(r, c, v)
    cc.font = Font(bold=bold, color=fg, size=size)
    cc.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bg: cc.fill = F(bg)
    if border: cc.border = Bd()
    return cc

# ── 차이원인 데이터 로드
wb_src = openpyxl.load_workbook(SRC, data_only=True)

def read_sheet(name):
    ws = wb_src[name]
    hdrs = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
    rows = []
    for r in range(3, ws.max_row+1):
        rv = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        if any(v is not None for v in rv):
            rows.append(dict(zip(hdrs, rv)))
    df = pd.DataFrame(rows)
    for col in ['계획금액','실적금액','차이(억)']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df

df_o = read_sheet('수주_차이원인')
df_s = read_sheet('매출_차이원인')

def gen_reason_text(df, top_n=8):
    """국내/해외별 차이 텍스트 생성"""
    lines = []
    for gh in ['국내', '해외']:
        sub = df[df['국내해외'] == gh].copy()
        if sub.empty: continue
        total = round(sub['차이(억)'].sum(), 1)
        gh_short = '국' if gh == '국내' else '해'
        sign = f'+{total:.0f}' if total >= 0 else f'{total:.0f}'
        lines.append(f'({gh_short} {sign}억)')

        # 소멸 (계획에만)
        extinct = sub[sub['원인분류'].str.startswith('소멸')].sort_values('차이(억)')
        for _, r in extinct.head(4).iterrows():
            proj = str(r['프로젝트_계획'] or '')[:20]
            if proj and proj != 'nan':
                lines.append(f"  소멸) {proj} {r['차이(억)']:+.1f}억")

        # 신규 (실적에만)
        new_ = sub[sub['원인분류'] == '신규'].sort_values('차이(억)', ascending=False)
        for _, r in new_.head(4).iterrows():
            proj = str(r['프로젝트_실적'] or '')[:20]
            if proj and proj != 'nan':
                lines.append(f"  신규) {proj} +{r['실적금액']:.1f}억")

        # 금액변동 (상위)
        chg = sub[sub['원인분류'].isin(['금액변동','명칭변경추정','분할추정'])]
        chg = chg[chg['차이(억)'].abs() >= 1].sort_values('차이(억)', key=abs, ascending=False)
        for _, r in chg.head(3).iterrows():
            proj = str(r['프로젝트_계획'] or r.get('프로젝트_실적','') or '')[:20]
            if proj and proj != 'nan':
                lines.append(f"  변동) {proj} {r['차이(억)']:+.1f}억")
    return '\n'.join(lines)

reason_order = gen_reason_text(df_o)
reason_sales = gen_reason_text(df_s)

ord_diff  = round(df_o['차이(억)'].sum())
sal_diff  = round(df_s['차이(억)'].sum())

ord_tr_diff = round(df_o[df_o['제품군']=='변압기']['차이(억)'].sum())
ord_cb_diff = round(df_o[df_o['제품군']=='차단기']['차이(억)'].sum())
sal_tr_diff = round(df_s[df_s['제품군']=='변압기']['차이(억)'].sum())
sal_cb_diff = round(df_s[df_s['제품군']=='차단기']['차이(억)'].sum())

# 변압기/차단기 따로 텍스트
def gen_prod_text(df):
    lines = []
    for prod in ['변압기', '차단기']:
        sub = df[df['제품군'] == prod]
        tot = round(sub['차이(억)'].sum(), 1)
        short = '변' if prod == '변압기' else '차'
        lines.append(f'({short} {tot:+.1f}억)')
        new_ = sub[sub['원인분류']=='신규'].sort_values('차이(억)', ascending=False).head(3)
        for _, r in new_.iterrows():
            proj = str(r['프로젝트_실적'] or '')[:18]
            if proj and proj != 'nan': lines.append(f'  신규) {proj} +{r["실적금액"]:.1f}억')
        ext = sub[sub['원인분류'].str.startswith('소멸')].sort_values('차이(억)').head(3)
        for _, r in ext.iterrows():
            proj = str(r['프로젝트_계획'] or '')[:18]
            if proj and proj != 'nan': lines.append(f'  소멸) {proj} {r["차이(억)"]:+.1f}억')
    return '\n'.join(lines)

prod_ord_text = gen_prod_text(df_o)
prod_sal_text = gen_prod_text(df_s)

# ══════════════════════════════════════════════════════════════
# Excel 생성
# ══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '1분기_차이분석'
ws.sheet_view.showGridLines = False

# 컬럼 레이아웃
# A=구분, B=항목, C-E=1분기(계획/실적/전년), F-H=4월, I-K=4월누계,
# L-N=5월, O-Q=6월, R-U=2분기, V-Y=상반기, Z-AA=7월, AB-AD=7월누계
col_widths = {
    'A':8, 'B':8,
    'C':9,'D':9,'E':9,          # 1분기
    'F':9,'G':9,'H':9,          # 4월
    'I':9,'J':9,'K':9,          # 4월누계
    'L':9,'M':9,'N':9,          # 5월
    'O':9,'P':9,'Q':9,          # 6월
    'R':9,'S':9,'T':9,'U':9,    # 2분기
    'V':9,'W':9,'X':9,'Y':9,    # 상반기
    'Z':9,'AA':9,               # 7월
    'AB':9,'AC':9,'AD':9,       # 7월 누계
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

LAST_COL = 30  # AD

# ── 타이틀
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_COL)
tc = ws.cell(1, 1, '■ 월별 수주/매출/이익 차이분석')
tc.font = Font(bold=True, size=11)
tc.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 18

# ── 헤더 행1: 기간명 (merged)
periods = [
    ('구분', 1, 2), ('', 1, 2),         # A,B 합치기
    ('1분기', 3, 5),
    ('4월', 6, 8),
    ('4월누계', 9, 11),
    ('5월', 12, 14),
    ('6월', 15, 17),
    ('2분기', 18, 21),
    ('상반기', 22, 25),
    ('7월', 26, 27),
    ('7월 누계', 28, 30),
]
for name, sc, ec in periods:
    if sc == ec:
        c = ws.cell(2, sc, name)
    else:
        ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        c = ws.cell(2, sc, name)
    c.font = Font(bold=True, size=9)
    c.fill = F('1F3864'); c.font = Font(bold=True, color='FFFFFF', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = Bd()
ws.row_dimensions[2].height = 16

# 구분/항목 merge 헤더
ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
for ci in (1, 2):
    cc = ws.cell(2, ci)
    cc.fill = F('1F3864'); cc.font = Font(bold=True, color='FFFFFF', size=9)
    cc.alignment = Alignment(horizontal='center', vertical='center')
    cc.border = Bd()

# ── 헤더 행2: 서브 (계획/실적/전년 등)
sub_headers = [
    # col: label
    (3,'계획'),(4,'실적'),(5,'전년'),
    (6,'계획'),(7,'보고'),(8,'실적'),
    (9,'계획'),(10,'보고'),(11,'실적'),
    (12,'계획'),(13,'보고'),(14,'예상'),
    (15,'계획'),(16,'보고'),(17,'예상'),
    (18,'계획'),(19,'보고'),(20,'예상'),(21,'전년'),
    (22,'계획'),(23,'보고'),(24,'예상'),(25,'전년'),
    (26,'계획'),(27,'예상'),
    (28,'계획'),(29,'예상'),(30,'전년'),
]
for ci, lbl in sub_headers:
    c = ws.cell(3, ci, lbl)
    c.font = Font(bold=True, size=8)
    c.fill = F(HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = Bd()
ws.row_dimensions[3].height = 14

# ── 데이터 행 (구분/항목/수치)
# 보고서 제공 수치 사용 (1분기)
report_data = {
    ('중전기','수주'):     {'계획':2607.5, '실적':4290.1, '전년':1864.0},
    ('중전기','매출'):     {'계획':1140.5, '실적':1214.3, '전년': 357.1},
    ('중전기','영업이익'): {'계획': 224.9, '실적': 293.2, '전년': 165.3},
    ('변압기','수주'):     {'계획':2231.4, '실적':3875.0, '전년':1464.9},
    ('변압기','매출'):     {'계획': 906.7, '실적': 974.2, '전년': 182.7},
    ('변압기','영업이익'): {'계획': 195.0, '실적': 247.7, '전년': 149.6},
    ('차단기','수주'):     {'계획': 376.2, '실적': 415.0, '전년': 399.2},
    ('차단기','매출'):     {'계획': 233.8, '실적': 240.1, '전년': 174.3},
    ('차단기','영업이익'): {'계획':  29.9, '실적':  45.5, '전년':  15.7},
}
PROD_BG = {'중전기':'DEEAF1','변압기':'E2EFDA','차단기':'FFF2CC'}
groups = [('중전기',['수주','매출','영업이익']),
          ('변압기',['수주','매출','영업이익']),
          ('차단기',['수주','매출','영업이익'])]

data_start = 4
r = data_start
for prod, items in groups:
    bg = PROD_BG[prod]
    row_start = r
    for item in items:
        nums = report_data.get((prod, item), {})
        # A: 구분 (나중에 merge)
        ws.cell(r, 1, '')
        ws.cell(r, 2, item).font = Font(size=9)
        ws.cell(r, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(r, 2).border = Bd()
        ws.cell(r, 2).fill = F(bg)
        # 1분기 수치
        for ci, key in [(3,'계획'),(4,'실적'),(5,'전년')]:
            v = nums.get(key, '')
            cc = ws.cell(r, ci, v)
            cc.font = Font(size=9)
            cc.alignment = Alignment(horizontal='center', vertical='center')
            cc.border = Bd()
            cc.fill = F(bg)
        # 나머지 빈 셀
        for ci in range(6, LAST_COL+1):
            cc = ws.cell(r, ci, '')
            cc.border = Bd()
        ws.row_dimensions[r].height = 14
        r += 1
    # 구분 셀 병합
    ws.merge_cells(start_row=row_start, start_column=1,
                   end_row=row_start+2, end_column=1)
    cc = ws.cell(row_start, 1, prod)
    cc.font = Font(bold=True, size=9)
    cc.fill = F(bg)
    cc.alignment = Alignment(horizontal='center', vertical='center')
    cc.border = Bd()

# 빈 행
ws.row_dimensions[r].height = 8
r += 1

# ══════════════════════════════════════════════════════════════
# 하단: 계획대비 차이분석
# ══════════════════════════════════════════════════════════════
# 섹션 제목
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST_COL)
tc = ws.cell(r, 1, '* 계획대비 차이분석 *')
tc.font = Font(bold=True, size=10)
tc.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r].height = 16
r += 1

# 차이분석 컬럼 레이아웃
# A(2): 기간, B-I(8): 수주, J-Q(8): 매출, R-Y(8): 영업이익, Z-AD(5): 변압기/차단기따로
# 각 영역: 처음 셀에 숫자, 나머지는 텍스트

DIFF_COLS = {
    '기간':    (1, 2),
    '수주':    (3, 10),
    '매출':    (11, 18),
    '영업이익': (19, 26),
    '변압기/차단기': (27, 30),
}

# 차이분석 헤더
for name, (sc, ec) in DIFF_COLS.items():
    ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=ec)
    cc = ws.cell(r, sc, name)
    cc.font = Font(bold=True, color='FFFFFF', size=9)
    cc.fill = F('1F3864')
    cc.alignment = Alignment(horizontal='center', vertical='center')
    cc.border = Bd()
ws.row_dimensions[r].height = 16
r += 1

# 차이분석 데이터 행
# 각 기간: 숫자행(2행) + 텍스트행(가변)
diff_rows = [
    # (기간라벨, 수주diff, 매출diff, 영익diff, 수주텍스트, 매출텍스트, 영익텍스트, 변압기차단기텍스트, n_text_rows)
    ('1분기', ord_diff, sal_diff, '', reason_order, reason_sales, '', prod_ord_text, 12),
    ('4월',   '',       '',       '', '', '', '', '', 4),
    ('4월누계','',       '',       '', '', '', '', '', 4),
    ('5월',   '',       '',       '', '', '', '', '', 4),
    ('6월',   '',       '',       '', '', '', '', '', 4),
    ('2분기', '',       '',       '', '', '', '', '', 4),
    ('상반기','',       '',       '', '', '', '', '', 4),
    ('7월',   '',       '',       '', '', '', '', '', 4),
    ('7월 누계','',     '',       '', '', '', '', '', 4),
]

for (period, ord_n, sal_n, oi_n,
     ord_txt, sal_txt, oi_txt, prod_txt, n_rows) in diff_rows:

    row_start = r
    total_rows = 1 + n_rows  # 숫자행 1 + 텍스트행

    # 기간 라벨 셀 (세로 merge)
    ws.merge_cells(start_row=r, start_column=1, end_row=r+total_rows-1, end_column=2)
    cc = ws.cell(r, 1, period)
    cc.font = Font(bold=True, size=9)
    cc.alignment = Alignment(horizontal='center', vertical='center')
    cc.border = Bd()

    # 숫자행 (노란 배경)
    for col_name, (sc, ec), num_v in [
        ('수주', DIFF_COLS['수주'], ord_n),
        ('매출', DIFF_COLS['매출'], sal_n),
        ('영업이익', DIFF_COLS['영업이익'], oi_n),
        ('변압기/차단기', DIFF_COLS['변압기/차단기'], ''),
    ]:
        ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=ec)
        cc = ws.cell(r, sc, num_v if num_v != '' else '')
        cc.font = Font(bold=True, size=11)
        cc.fill = F(YELLOW)
        cc.alignment = Alignment(horizontal='center', vertical='center')
        cc.border = Bd()
    ws.row_dimensions[r].height = 18
    r += 1

    # 텍스트행 (내용 + 빈 행들)
    text_data = {
        DIFF_COLS['수주']:          ord_txt,
        DIFF_COLS['매출']:          sal_txt,
        DIFF_COLS['영업이익']:      oi_txt,
        DIFF_COLS['변압기/차단기']: prod_txt,
    }
    # 텍스트 전체를 하나의 merged cell에 넣기
    for (sc, ec), txt in text_data.items():
        ws.merge_cells(start_row=r, start_column=sc,
                       end_row=r + n_rows - 1, end_column=ec)
        cc = ws.cell(r, sc, txt)
        cc.font = Font(size=8)
        cc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cc.border = Bd()
    # 기간 라벨 아래 빈 셀들도 border
    for ri in range(r, r + n_rows):
        ws.row_dimensions[ri].height = 13
        for ci in (1, 2):
            ws.cell(ri, ci).border = Bd()

    r += n_rows

wb.save(OUT)
print(f'저장 완료: {OUT}')
