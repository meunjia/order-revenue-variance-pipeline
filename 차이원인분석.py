#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""차이원인분석 — 수주/매출 소멸·신규·변동 번호 정리"""

import os, sys, io, openpyxl, pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

WORK_DIR = r"C:\Users\user\claudecode\월별차이"
SRC      = os.path.join(WORK_DIR, "차이분석_결과.xlsx")
OUT      = os.path.join(WORK_DIR, "차이원인분석.xlsx")

# ── 원본 데이터 로드
src = openpyxl.load_workbook(SRC, read_only=True)

def load_sheet(name):
    ws = src[name]
    rows = list(ws.iter_rows(values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0])

m_ord = load_sheet('수주_프로젝트심층')
m_sal = load_sheet('매출_프로젝트심층')
ne    = load_sheet('신규_소멸_프로젝트')
src.close()

ord_소멸 = ne[(ne['구분']=='수주') & ne['분류'].str.contains('소멸', na=False)].sort_values('금액', ascending=False).reset_index(drop=True)
ord_신규 = ne[(ne['구분']=='수주') & ne['분류'].str.contains('신규', na=False)].sort_values('금액', ascending=False).reset_index(drop=True)
sal_소멸 = ne[(ne['구분']=='매출') & ne['분류'].str.contains('소멸', na=False)].sort_values('금액', ascending=False).reset_index(drop=True)
sal_신규 = ne[(ne['구분']=='매출') & ne['분류'].str.contains('신규', na=False)].sort_values('금액', ascending=False).reset_index(drop=True)
m_ord_변동 = m_ord.copy().sort_values('차이(억)', key=abs, ascending=False).reset_index(drop=True)
m_sal_변동 = m_sal.copy().sort_values('차이(억)', key=abs, ascending=False).reset_index(drop=True)

# ── 집계
계획_수주 = 2607.5; 실적_수주 = 4290.1
계획_매출 = 1140.5; 실적_매출 = 1214.3

신규합_수주  = round(ord_신규['금액'].sum(), 1)
소멸합_수주  = round(ord_소멸['금액'].sum(), 1)
변동합_수주  = round(m_ord_변동['차이(억)'].sum(), 1)
신규합_매출  = round(sal_신규['금액'].sum(), 1)
소멸합_매출  = round(sal_소멸['금액'].sum(), 1)
변동합_매출  = round(m_sal_변동['차이(억)'].sum(), 1)

# ── 스타일
T = Side(style='thin')
M = Side(style='medium')
BD   = Border(left=T, right=T, top=T, bottom=T)
BDL  = Border(left=M, right=T, top=T, bottom=T)

def fill(h): return PatternFill('solid', fgColor=h)
BLU  = fill('4472C4')
LBLU = fill('D6E4F7')
GRN  = fill('E2EFDA')
RED  = fill('FFDDE1')
YEL  = fill('FFFF00')
GRY  = fill('F2F2F2')
WHT  = fill('FFFFFF')
ORG  = fill('FCE4D6')

def hf(sz=9):  return Font(bold=True, size=sz, color='FFFFFF', name='맑은 고딕')
def bf(sz=9):  return Font(bold=True, size=sz, color='000000', name='맑은 고딕')
def df(sz=9):  return Font(bold=False, size=sz, color='000000', name='맑은 고딕')
def C(w=False): return Alignment(horizontal='center', vertical='center', wrap_text=w)
def L(w=True):  return Alignment(horizontal='left',   vertical='center', wrap_text=w)

def write_row(ws, row, vals_styles, height=15):
    ws.row_dimensions[row].height = height
    for col, (val, font, fill_, align, fmt) in enumerate(vals_styles, 1):
        c = ws.cell(row, col, value=val)
        c.font  = font
        c.fill  = fill_
        c.alignment = align
        c.border = BD
        if fmt: c.number_format = fmt

def header_row(ws, row, headers, fills, height=15):
    ws.row_dimensions[row].height = height
    for col, (h, f) in enumerate(zip(headers, fills), 1):
        c = ws.cell(row, col, value=h)
        c.font  = hf()
        c.fill  = f
        c.alignment = C()
        c.border = BD

def section_title(ws, row, text, ncols, fill_=None):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws['A' + str(row)]
    c.value = text
    c.font  = bf(10)
    c.fill  = fill_ or GRY
    c.alignment = L(False)
    c.border = BD
    ws.row_dimensions[row].height = 18

# ══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════
# 시트 1: 요약 (흐름)
# ══════════════════════════════════════════════════════════════
ws = wb.create_sheet('요약_흐름')
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 16

# 타이틀
ws.merge_cells('A1:F1')
c = ws['A1']
c.value = '1분기 수주/매출 차이 원인 — 흐름 요약'
c.font  = bf(12)
c.alignment = C()
ws.row_dimensions[1].height = 24

# 헤더
header_row(ws, 2,
    ['구분', '계획(억)', '신규(+)', '소멸(-)', '변동(±)', '실적(억)'],
    [BLU,    BLU,       BLU,      BLU,      BLU,      BLU],
    height=16)

# 데이터
rows_sum = [
    ('수주', 계획_수주, f'+{신규합_수주:,.1f}', f'-{소멸합_수주:,.1f}', f'{변동합_수주:+,.1f}', 실적_수주),
    ('매출', 계획_매출, f'+{신규합_매출:,.1f}', f'-{소멸합_매출:,.1f}', f'{변동합_매출:+,.1f}', f'{실적_매출} *'),
]
for r_idx, (구분, 계획, 신규, 소멸, 변동, 실적) in enumerate(rows_sum, 3):
    ws.row_dimensions[r_idx].height = 18
    for col, (val, fnt, fl_, aln) in enumerate([
        (구분, bf(10), LBLU, C()),
        (계획, bf(10), WHT,  C()),
        (신규, bf(10), GRN,  C()),
        (소멸, bf(10), RED,  C()),
        (변동, bf(10), ORG,  C()),
        (실적, bf(10), YEL,  C()),
    ], 1):
        c = ws.cell(r_idx, col, value=val)
        c.font = fnt; c.fill = fl_; c.alignment = aln; c.border = BD

# 비고
ws.merge_cells('A5:F5')
c = ws['A5']
c.value = '* 매출 실적 1,214.3억은 보고서 기준. CSV 계산값 1,186.8억 (USD 환율 1,400 vs 1,300 차이 -27.5억)'
c.font  = df(8); c.fill = fill('FFF2CC'); c.alignment = L()

# 수식 설명
ws.merge_cells('A7:F7')
ws['A7'].value = '▶ 계산 구조 :  실적 = 계획  +  신규  −  소멸  ±  변동 (금액 변동된 매칭 프로젝트)'
ws['A7'].font  = bf(9); ws['A7'].fill = LBLU; ws['A7'].alignment = L()
ws.row_dimensions[7].height = 16

# 소계 박스 (수주)
ws.merge_cells('A9:F9')
ws['A9'].value = f'수주: {계획_수주:,.1f}  +  {신규합_수주:,.1f}  −  {소멸합_수주:,.1f}  {변동합_수주:+,.1f}  =  {실적_수주:,.1f}  (차이 {실적_수주-계획_수주:+,.1f}억)'
ws['A9'].font  = bf(10); ws['A9'].fill = GRN; ws['A9'].alignment = C()
ws.row_dimensions[9].height = 18

ws.merge_cells('A10:F10')
ws['A10'].value = f'매출: {계획_매출:,.1f}  +  {신규합_매출:,.1f}  −  {소멸합_매출:,.1f}  {변동합_매출:+,.1f}  =  {계획_매출+신규합_매출-소멸합_매출+변동합_매출:,.1f}  (CSV기준. 보고서 {실적_매출:,.1f}, 차이 {실적_매출-계획_매출:+,.1f}억)'
ws['A10'].font  = bf(10); ws['A10'].fill = ORG; ws['A10'].alignment = C()
ws.row_dimensions[10].height = 18
for r in range(9, 11):
    ws.cell(r, 1).border = BD

# ══════════════════════════════════════════════════════════════
# 시트 2: 수주 상세
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('수주_상세')
ws2.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [5,18,8,8,10,10,10,10,20]):
    ws2.column_dimensions[col].width = w

# 타이틀
ws2.merge_cells('A1:I1')
c = ws2['A1']
c.value = f'수주 차이 상세  |  계획 {계획_수주:,.1f}억 → 실적 {실적_수주:,.1f}억  (차이 {실적_수주-계획_수주:+,.1f}억)'
c.font = bf(11); c.alignment = C(); c.fill = LBLU; ws2.row_dimensions[1].height = 22

cur_row = 2

# ─ 소멸
section_title(ws2, cur_row,
    f'① 소멸/이월  {len(ord_소멸)}건  합계 -{소멸합_수주:,.1f}억  (계획에 있었으나 실적에 없음)', 9, RED)
cur_row += 1
header_row(ws2, cur_row,
    ['No','프로젝트명','제품군','국내해외','팀명','계획금액(억)','분류','','비고'],
    [BLU]*9)
cur_row += 1
for i, r in ord_소멸.iterrows():
    write_row(ws2, cur_row, [
        (i+1,                bf(),  WHT, C(), None),
        (r['프로젝트'],       df(),  WHT, L(),  None),
        (r['제품군'],         df(),  WHT, C(), None),
        (r['국내해외'],       df(),  WHT, C(), None),
        (r['팀명'],           df(),  WHT, C(), None),
        (-r['금액'],          bf(),  RED, C(), '#,##0.0'),
        (r['분류'],           df(8), WHT, L(),  None),
        ('', df(), WHT, C(), None),
        ('', df(), WHT, C(), None),
    ])
    cur_row += 1

# 소멸 소계
ws2.merge_cells(f'A{cur_row}:E{cur_row}')
ws2[f'A{cur_row}'].value = '소멸 소계'
ws2[f'A{cur_row}'].font = bf(); ws2[f'A{cur_row}'].fill = RED; ws2[f'A{cur_row}'].alignment = C()
ws2.cell(cur_row, 6, value=-소멸합_수주).font = bf()
ws2.cell(cur_row, 6).fill = RED; ws2.cell(cur_row, 6).alignment = C()
ws2.cell(cur_row, 6).number_format = '#,##0.0'
ws2.cell(cur_row, 6).border = BD
for col in [1,2,3,4,5]: ws2.cell(cur_row, col).border = BD
ws2.row_dimensions[cur_row].height = 15
cur_row += 2

# ─ 신규
section_title(ws2, cur_row,
    f'② 신규  {len(ord_신규)}건  합계 +{신규합_수주:,.1f}억  (실적에 있으나 계획에 없음)', 9, GRN)
cur_row += 1
header_row(ws2, cur_row,
    ['No','프로젝트명','제품군','국내해외','팀명','실적금액(억)','분류','','비고'],
    [BLU]*9)
cur_row += 1
for i, r in ord_신규.iterrows():
    write_row(ws2, cur_row, [
        (i+1,          bf(),  WHT, C(), None),
        (r['프로젝트'], df(),  WHT, L(),  None),
        (r['제품군'],   df(),  WHT, C(), None),
        (r['국내해외'], df(),  WHT, C(), None),
        (r['팀명'],     df(),  WHT, C(), None),
        (r['금액'],     bf(),  GRN, C(), '#,##0.0'),
        (r['분류'],     df(8), WHT, L(),  None),
        ('', df(), WHT, C(), None),
        ('', df(), WHT, C(), None),
    ])
    cur_row += 1

# 신규 소계
ws2.merge_cells(f'A{cur_row}:E{cur_row}')
ws2[f'A{cur_row}'].value = '신규 소계'
ws2[f'A{cur_row}'].font = bf(); ws2[f'A{cur_row}'].fill = GRN; ws2[f'A{cur_row}'].alignment = C()
ws2.cell(cur_row, 6, value=신규합_수주).font = bf()
ws2.cell(cur_row, 6).fill = GRN; ws2.cell(cur_row, 6).alignment = C()
ws2.cell(cur_row, 6).number_format = '#,##0.0'
ws2.cell(cur_row, 6).border = BD
for col in [1,2,3,4,5]: ws2.cell(cur_row, col).border = BD
ws2.row_dimensions[cur_row].height = 15
cur_row += 2

# ─ 변동
section_title(ws2, cur_row,
    f'③ 금액 변동  {len(m_ord_변동)}건  합계 {변동합_수주:+,.1f}억  (계획↔실적 매칭, 금액 차이)', 9, ORG)
cur_row += 1
header_row(ws2, cur_row,
    ['No','프로젝트명(계획)','제품군','국내해외','팀명','계획금액(억)','실적금액(억)','차이(억)','매칭유형'],
    [BLU]*9)
cur_row += 1
for i, r in m_ord_변동.iterrows():
    diff = r['차이(억)']
    diff_fill = GRN if diff > 0 else (RED if diff < 0 else WHT)
    write_row(ws2, cur_row, [
        (i+1,                     bf(),  WHT,       C(), None),
        (r['프로젝트_계획'],        df(),  WHT,       L(),  None),
        (r['제품군'],              df(),  WHT,       C(), None),
        (r['국내해외'],            df(),  WHT,       C(), None),
        (r['팀명'],                df(),  WHT,       C(), None),
        (r['계획금액'],            df(),  WHT,       C(), '#,##0.0'),
        (r['실적금액'],            df(),  WHT,       C(), '#,##0.0'),
        (diff,                    bf(),  diff_fill, C(), '+#,##0.0;-#,##0.0'),
        (r['매칭유형'],            df(8), WHT,       C(), None),
    ])
    cur_row += 1

# 변동 소계
ws2.merge_cells(f'A{cur_row}:G{cur_row}')
ws2[f'A{cur_row}'].value = '변동 소계'
ws2[f'A{cur_row}'].font = bf(); ws2[f'A{cur_row}'].fill = ORG; ws2[f'A{cur_row}'].alignment = C()
ws2.cell(cur_row, 8, value=변동합_수주).font = bf()
ws2.cell(cur_row, 8).fill = ORG; ws2.cell(cur_row, 8).alignment = C()
ws2.cell(cur_row, 8).number_format = '+#,##0.0;-#,##0.0'
ws2.cell(cur_row, 8).border = BD
for col in range(1,8): ws2.cell(cur_row, col).border = BD
ws2.row_dimensions[cur_row].height = 15
cur_row += 2

# 최종 합계
ws2.merge_cells(f'A{cur_row}:G{cur_row}')
ws2[f'A{cur_row}'].value = f'계획 {계획_수주:,.1f}  +  신규 {신규합_수주:,.1f}  −  소멸 {소멸합_수주:,.1f}  {변동합_수주:+,.1f}  =  실적 {실적_수주:,.1f}'
ws2[f'A{cur_row}'].font = bf(10); ws2[f'A{cur_row}'].fill = YEL; ws2[f'A{cur_row}'].alignment = C()
ws2.cell(cur_row, 8, value=실적_수주-계획_수주).font = bf(10)
ws2.cell(cur_row, 8).fill = YEL; ws2.cell(cur_row, 8).alignment = C()
ws2.cell(cur_row, 8).number_format = '+#,##0.0;-#,##0.0'
ws2.cell(cur_row, 8).border = BD
for col in range(1,8): ws2.cell(cur_row, col).border = BD
ws2.row_dimensions[cur_row].height = 20

# ══════════════════════════════════════════════════════════════
# 시트 3: 매출 상세 (동일 구조)
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('매출_상세')
ws3.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [5,18,8,8,10,10,10,10,20]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells('A1:I1')
c = ws3['A1']
c.value = f'매출 차이 상세  |  계획 {계획_매출:,.1f}억 → 실적 {실적_매출:,.1f}억  (차이 {실적_매출-계획_매출:+,.1f}억, CSV기준 {계획_매출+신규합_매출-소멸합_매출+변동합_매출:,.1f}억)'
c.font = bf(11); c.alignment = C(); c.fill = LBLU; ws3.row_dimensions[1].height = 22

cur_row = 2

for 구분, data, 소멸df, 신규df, 소멸합, 신규합, 변동합, 변동df, 계획, 실적, sect_no in [
    ('매출', None, sal_소멸, sal_신규, 소멸합_매출, 신규합_매출, 변동합_매출, m_sal_변동, 계획_매출, 실적_매출, '①②③'),
]:
    # 소멸
    section_title(ws3, cur_row,
        f'① 소멸/이월  {len(소멸df)}건  합계 -{소멸합:,.1f}억', 9, RED)
    cur_row += 1
    header_row(ws3, cur_row,
        ['No','프로젝트명','제품군','국내해외','팀명','계획금액(억)','분류','','비고'], [BLU]*9)
    cur_row += 1
    for i, r in 소멸df.iterrows():
        write_row(ws3, cur_row, [
            (i+1,           bf(),  WHT, C(), None),
            (r['프로젝트'],  df(),  WHT, L(),  None),
            (r['제품군'],    df(),  WHT, C(), None),
            (r['국내해외'],  df(),  WHT, C(), None),
            (r['팀명'],      df(),  WHT, C(), None),
            (-r['금액'],     bf(),  RED, C(), '#,##0.0'),
            (r['분류'],      df(8), WHT, L(),  None),
            ('', df(), WHT, C(), None), ('', df(), WHT, C(), None),
        ])
        cur_row += 1
    ws3.merge_cells(f'A{cur_row}:E{cur_row}')
    ws3[f'A{cur_row}'].value = '소멸 소계'; ws3[f'A{cur_row}'].font = bf()
    ws3[f'A{cur_row}'].fill = RED; ws3[f'A{cur_row}'].alignment = C()
    ws3.cell(cur_row,6, value=-소멸합).number_format='#,##0.0'
    ws3.cell(cur_row,6).font=bf(); ws3.cell(cur_row,6).fill=RED
    ws3.cell(cur_row,6).alignment=C(); ws3.cell(cur_row,6).border=BD
    for col in range(1,6): ws3.cell(cur_row,col).border=BD
    ws3.row_dimensions[cur_row].height=15; cur_row += 2

    # 신규
    section_title(ws3, cur_row,
        f'② 신규  {len(신규df)}건  합계 +{신규합:,.1f}억', 9, GRN)
    cur_row += 1
    header_row(ws3, cur_row,
        ['No','프로젝트명','제품군','국내해외','팀명','실적금액(억)','분류','','비고'], [BLU]*9)
    cur_row += 1
    for i, r in 신규df.iterrows():
        write_row(ws3, cur_row, [
            (i+1,           bf(),  WHT, C(), None),
            (r['프로젝트'],  df(),  WHT, L(),  None),
            (r['제품군'],    df(),  WHT, C(), None),
            (r['국내해외'],  df(),  WHT, C(), None),
            (r['팀명'],      df(),  WHT, C(), None),
            (r['금액'],      bf(),  GRN, C(), '#,##0.0'),
            (r['분류'],      df(8), WHT, L(),  None),
            ('', df(), WHT, C(), None), ('', df(), WHT, C(), None),
        ])
        cur_row += 1
    ws3.merge_cells(f'A{cur_row}:E{cur_row}')
    ws3[f'A{cur_row}'].value = '신규 소계'; ws3[f'A{cur_row}'].font = bf()
    ws3[f'A{cur_row}'].fill = GRN; ws3[f'A{cur_row}'].alignment = C()
    ws3.cell(cur_row,6, value=신규합).number_format='#,##0.0'
    ws3.cell(cur_row,6).font=bf(); ws3.cell(cur_row,6).fill=GRN
    ws3.cell(cur_row,6).alignment=C(); ws3.cell(cur_row,6).border=BD
    for col in range(1,6): ws3.cell(cur_row,col).border=BD
    ws3.row_dimensions[cur_row].height=15; cur_row += 2

    # 변동
    section_title(ws3, cur_row,
        f'③ 금액 변동  {len(변동df)}건  합계 {변동합:+,.1f}억', 9, ORG)
    cur_row += 1
    header_row(ws3, cur_row,
        ['No','프로젝트명(계획)','제품군','국내해외','팀명','계획금액(억)','실적금액(억)','차이(억)','매칭유형'],
        [BLU]*9)
    cur_row += 1
    for i, r in 변동df.iterrows():
        diff = r['차이(억)']
        diff_fill = GRN if diff > 0 else (RED if diff < 0 else WHT)
        write_row(ws3, cur_row, [
            (i+1,                  bf(),  WHT,       C(), None),
            (r['프로젝트_계획'],    df(),  WHT,       L(),  None),
            (r['제품군'],          df(),  WHT,       C(), None),
            (r['국내해외'],        df(),  WHT,       C(), None),
            (r['팀명'],            df(),  WHT,       C(), None),
            (r['계획금액'],        df(),  WHT,       C(), '#,##0.0'),
            (r['실적금액'],        df(),  WHT,       C(), '#,##0.0'),
            (diff,                bf(),  diff_fill, C(), '+#,##0.0;-#,##0.0'),
            (r['매칭유형'],        df(8), WHT,       C(), None),
        ])
        cur_row += 1
    ws3.merge_cells(f'A{cur_row}:G{cur_row}')
    ws3[f'A{cur_row}'].value='변동 소계'; ws3[f'A{cur_row}'].font=bf()
    ws3[f'A{cur_row}'].fill=ORG; ws3[f'A{cur_row}'].alignment=C()
    ws3.cell(cur_row,8,value=변동합).number_format='+#,##0.0;-#,##0.0'
    ws3.cell(cur_row,8).font=bf(); ws3.cell(cur_row,8).fill=ORG
    ws3.cell(cur_row,8).alignment=C(); ws3.cell(cur_row,8).border=BD
    for col in range(1,8): ws3.cell(cur_row,col).border=BD
    ws3.row_dimensions[cur_row].height=15; cur_row += 2

    csv_실적 = round(계획+신규합-소멸합+변동합, 1)
    ws3.merge_cells(f'A{cur_row}:G{cur_row}')
    ws3[f'A{cur_row}'].value = f'계획 {계획:,.1f}  +  신규 {신규합:,.1f}  −  소멸 {소멸합:,.1f}  {변동합:+,.1f}  =  {csv_실적:,.1f}  (보고서 {실적:,.1f}, 차이 -27.5억 = 환율)'
    ws3[f'A{cur_row}'].font=bf(10); ws3[f'A{cur_row}'].fill=YEL; ws3[f'A{cur_row}'].alignment=C()
    ws3.cell(cur_row,8,value=실적-계획).font=bf(10)
    ws3.cell(cur_row,8).fill=YEL; ws3.cell(cur_row,8).alignment=C()
    ws3.cell(cur_row,8).number_format='+#,##0.0;-#,##0.0'; ws3.cell(cur_row,8).border=BD
    for col in range(1,8): ws3.cell(cur_row,col).border=BD
    ws3.row_dimensions[cur_row].height=20

wb.save(OUT)
print(f'저장 완료: {OUT}')
print(f'  시트: 요약_흐름 / 수주_상세 / 매출_상세')
print(f'  수주: 소멸 {len(ord_소멸)}건, 신규 {len(ord_신규)}건, 변동 {len(m_ord_변동)}건')
print(f'  매출: 소멸 {len(sal_소멸)}건, 신규 {len(sal_신규)}건, 변동 {len(m_sal_변동)}건')
