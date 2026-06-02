#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""월별 수주/매출/이익 차이분석 서식 생성"""

import os, sys, io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

WORK_DIR = r"C:\Users\user\claudecode\월별차이"
OUT      = os.path.join(WORK_DIR, "월별차이분석_1분기.xlsx")

# ── 1. 분석 결과에서 텍스트 사유 읽기
src = openpyxl.load_workbook(os.path.join(WORK_DIR, "차이분석_결과.xlsx"), read_only=True)
ws_txt = src['텍스트_사유']
rows_txt = [(r[0], r[1]) for r in ws_txt.iter_rows(values_only=True)]
src.close()
ord_lines = [b for a, b in rows_txt if a == '수주' and b]
sal_lines = [b for a, b in rows_txt if a == '매출' and b]

# 수주: 소멸/신규 분리
ord_extinct = [l for l in ord_lines if '취소/이월' in l]
ord_new     = [l for l in ord_lines if '신규' in l]
# 매출: 소멸/신규 분리
sal_extinct = [l for l in sal_lines if '취소/이월' in l]
sal_new     = [l for l in sal_lines if '신규' in l]

def make_text(extinct, new, max_lines=18):
    lines = []
    if extinct:
        lines.append('■ 소멸/이월 (계획→미반영)')
        lines += extinct[:9]
    if new:
        lines.append('')
        lines.append('■ 신규 (계획외 수주)')
        lines += new[:8]
    return '\n'.join(lines[:max_lines])

ord_text = make_text(ord_extinct, ord_new)
sal_text = make_text(sal_extinct, sal_new)

# ── 2. 스타일 정의
def fill(hex_): return PatternFill('solid', fgColor=hex_)
T = Side(style='thin')
BD   = Border(left=T, right=T, top=T, bottom=T)
BD_L = Border(left=Side(style='medium'), right=T, top=T, bottom=T)

HDR_FILL  = fill('4472C4')  # 파란색 헤더
SUB_FILL  = fill('D6E4F7')  # 연파란색 서브헤더
YELL_FILL = fill('FFFF00')  # 노란색 차이
GRP1_FILL = fill('FFFFFF')  # 흰색
GRP2_FILL = fill('F5F8FF')  # 아주 연한 파란 (변압기)
GRP3_FILL = fill('FAFAFA')  # 연회색 (차단기)

def hf(size=9):  return Font(bold=True, size=size, color='FFFFFF', name='맑은 고딕')
def df(size=9):  return Font(bold=False, size=size, color='000000', name='맑은 고딕')
def bf(size=9):  return Font(bold=True,  size=size, color='000000', name='맑은 고딕')
def C(wrap=False): return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def TL(wrap=True): return Alignment(horizontal='left', vertical='top',    wrap_text=wrap)

# ── 3. 데이터
Q1 = {
    '중전기': {
        '수주':    {'계획': 2607.5, '실적': 4290.1, '전년': 1864.0},
        '매출':    {'계획': 1140.5, '실적': 1214.3, '전년':  357.1},
        '영업이익':{'계획':  224.9, '실적':  293.2, '전년':  165.3},
    },
    '변압기': {
        '수주':    {'계획': 2231.4, '실적': 3875.0, '전년': 1464.9},
        '매출':    {'계획':  906.7, '실적':  974.2, '전년':  182.7},
        '영업이익':{'계획':  195.0, '실적':  247.7, '전년':  149.6},
    },
    '차단기': {
        '수주':    {'계획':  376.2, '실적':  415.0, '전년':  399.2},
        '매출':    {'계획':  233.8, '실적':  240.1, '전년':  174.3},
        '영업이익':{'계획':   29.9, '실적':   45.5, '전년':   15.7},
    },
}

# 1분기 차이 (보고서 기준)
DIFF_Q1 = {
    '수주':    round(Q1['중전기']['수주']['실적']    - Q1['중전기']['수주']['계획']),
    '매출':    round(Q1['중전기']['매출']['실적']    - Q1['중전기']['매출']['계획']),
    '영업이익':round(Q1['중전기']['영업이익']['실적']- Q1['중전기']['영업이익']['계획']),
}

# 기간 컬럼 구조: (시작col_idx, 기간명, [서브헤더])
PERIODS = [
    (3,  '1분기',   ['계획', '실적', '전년']),
    (6,  '4월',     ['계획', '보고', '실적']),
    (9,  '4월누계', ['계획', '보고', '실적']),
    (12, '5월',     ['계획', '보고', '예상']),
    (15, '6월',     ['계획', '보고', '예상']),
    (18, '2분기',   ['계획', '보고', '예상', '전년']),
    (22, '상반기',  ['계획', '보고', '예상', '전년']),
    (26, '7월',     ['계획', '예상']),
    (28, '7월누계', ['계획', '예상', '전년']),
]
MAX_COL = 30  # A=1 ... AD=30

# ── 4. 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "월별차이분석"
ws.sheet_view.showGridLines = False

# 컬럼 너비
ws.column_dimensions['A'].width = 5.5
ws.column_dimensions['B'].width = 7.5
for i in range(3, MAX_COL + 1):
    ws.column_dimensions[get_column_letter(i)].width = 8.5

# ══════════════════════════════════════════
# [상단 테이블] rows 1~12
# ══════════════════════════════════════════

# Row 1: 타이틀
ws.merge_cells(f'A1:{get_column_letter(MAX_COL)}1')
c = ws['A1']
c.value = '■ 월별 수주/매출/이익 차이분석'
c.font = Font(bold=True, size=11, name='맑은 고딕')
c.alignment = C()
ws.row_dimensions[1].height = 22

# Row 2: 기간 헤더 / Row 3: 서브헤더
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 14

# 구분 셀 (A2:B3 병합)
ws.merge_cells('A2:B3')
c = ws['A2']
c.value = '구분'
c.font = hf()
c.fill = HDR_FILL
c.alignment = C()
c.border = BD

for start_col, period, subs in PERIODS:
    end_col = start_col + len(subs) - 1
    sc = get_column_letter(start_col)
    ec = get_column_letter(end_col)

    # 기간 헤더 (row 2, merged)
    ws.merge_cells(f'{sc}2:{ec}2')
    c = ws[f'{sc}2']
    c.value = period
    c.font = hf()
    c.fill = HDR_FILL
    c.alignment = C()
    for col in range(start_col, end_col + 1):
        ws.cell(2, col).border = BD

    # 서브헤더 (row 3)
    for j, sub in enumerate(subs):
        c = ws.cell(3, start_col + j)
        c.value = sub
        c.font = bf(size=8)
        c.fill = SUB_FILL
        c.alignment = C()
        c.border = BD

# Data rows (4~12)
groups      = ['중전기', '변압기', '차단기']
items       = ['수주', '매출', '영업이익']
grp_fills   = [GRP1_FILL, GRP2_FILL, GRP3_FILL]
grp_heights = [14, 14, 14]

for g, grp in enumerate(groups):
    base = 4 + g * 3
    grp_f = grp_fills[g]
    for r in range(base, base + 3):
        ws.row_dimensions[r].height = 14

    # Col A: 대분류 (병합)
    ws.merge_cells(f'A{base}:A{base+2}')
    c = ws[f'A{base}']
    c.value = grp
    c.font = bf(size=9)
    c.fill = grp_f
    c.alignment = C(wrap=True)
    c.border = BD

    for i, item in enumerate(items):
        row = base + i

        # Col B: 항목
        c = ws.cell(row, 2)
        c.value = item
        c.font = df(size=9)
        c.fill = grp_f
        c.alignment = C()
        c.border = BD

        # 1분기 데이터 (cols 3~5)
        vals = Q1[grp][item]
        for j, key in enumerate(['계획', '실적', '전년']):
            c = ws.cell(row, 3 + j)
            c.value = vals[key]
            c.font = df(size=9)
            c.fill = grp_f
            c.alignment = C()
            c.border = BD
            c.number_format = '#,##0.0'

        # 나머지 기간 (빈칸, 서식만)
        for start_col, _, subs in PERIODS[1:]:
            for j in range(len(subs)):
                c = ws.cell(row, start_col + j)
                c.fill = grp_f
                c.border = BD
                c.alignment = C()
                c.number_format = '#,##0.0'

# A열/B열 border fix for rows 2~3
for row in [2, 3]:
    for col in [1, 2]:
        ws.cell(row, col).border = BD

# ══════════════════════════════════════════
# [차이분석 섹션] rows 14~
# ══════════════════════════════════════════
# 차이분석 섹션 컬럼 배치:
#   A      : 기간
#   B~L(11): 수주
#   M~W(11): 매출
#   X~AD(7): 영업이익

DR = 14   # 차이분석 시작 행
ws.row_dimensions[13].height = 6   # 간격
ws.row_dimensions[DR].height = 16

# 타이틀
ws.merge_cells(f'A{DR}:{get_column_letter(MAX_COL)}{DR}')
c = ws[f'A{DR}']
c.value = '★ 계획대비 차이분석 ★'
c.font = bf(size=10)
c.alignment = Alignment(horizontal='left', vertical='center')

# 섹션 헤더 (row 15)
HR = DR + 1
ws.row_dimensions[HR].height = 16

# 기간 헤더
c = ws.cell(HR, 1)
c.value = '구분'
c.font = hf()
c.fill = HDR_FILL
c.alignment = C()
c.border = BD

S_COLS = {'수주': (2, 12), '매출': (13, 23), '영업이익': (24, 30)}
for label, (s, e) in S_COLS.items():
    ws.merge_cells(f'{get_column_letter(s)}{HR}:{get_column_letter(e)}{HR}')
    c = ws[f'{get_column_letter(s)}{HR}']
    c.value = label
    c.font = hf()
    c.fill = HDR_FILL
    c.alignment = C()
    for col in range(s, e + 1):
        ws.cell(HR, col).border = BD

# 기간 행들
DIFF_ROWS = [
    ('1분기', True),   # True = 데이터 있음
    ('4월',   False),
    ('4월누계',False),
    ('5월',   False),
    ('6월',   False),
    ('2분기', False),
    ('상반기', False),
    ('7월',   False),
    ('7월누계',False),
]

for idx, (period, has_data) in enumerate(DIFF_ROWS):
    row = HR + 1 + idx
    h = 150 if has_data else 55
    ws.row_dimensions[row].height = h

    # 기간 레이블
    c = ws.cell(row, 1)
    c.value = period
    c.font = bf(size=9)
    c.fill = SUB_FILL if not has_data else fill('E2EFDA')
    c.alignment = C(wrap=True)
    c.border = BD

    for label, (s, e) in S_COLS.items():
        diff_val = DIFF_Q1.get(label, None) if has_data else None

        # 숫자 + 텍스트 조합
        if has_data and diff_val is not None:
            sign = '+' if diff_val >= 0 else ''
            if label == '수주':
                body = ord_text
            elif label == '매출':
                body = sal_text
            else:
                body = ''
            cell_val = f'{sign}{diff_val:,}\n\n{body}'
        else:
            cell_val = ''

        ws.merge_cells(f'{get_column_letter(s)}{row}:{get_column_letter(e)}{row}')
        c = ws[f'{get_column_letter(s)}{row}']
        c.value = cell_val
        c.font = df(size=8)
        c.fill = YELL_FILL if has_data else GRP1_FILL
        c.alignment = TL()
        for col in range(s, e + 1):
            ws.cell(row, col).border = BD

# ── 5. 저장
wb.save(OUT)
print(f'저장 완료: {OUT}')
