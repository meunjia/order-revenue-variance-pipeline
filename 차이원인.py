#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""수주/매출 계획 vs 실적 차이 원인 분석"""

import os, re, sys, io, warnings
import pandas as pd
import numpy as np
import chardet
from rapidfuzz import fuzz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

WORK_DIR = r"C:\Users\user\claudecode\월별차이"
CSV_DIR  = os.path.join(WORK_DIR, "csv")
OUT      = os.path.join(WORK_DIR, "차이원인분석.xlsx")

Q1_PLAN   = ['01월', '02월', '03월']
Q1_ACTUAL = ['26년01월', '26년02월', '26년03월']

# ============================================================
# 유틸
# ============================================================
def detect_enc(path):
    with open(path, 'rb') as f:
        return chardet.detect(f.read(100000))['encoding'] or 'utf-8-sig'

def load_csv(fname):
    path = os.path.join(CSV_DIR, fname)
    df = pd.read_csv(path, encoding=detect_enc(path), low_memory=False)
    df.columns = [str(c).strip().replace('\n', '') for c in df.columns]
    df.dropna(how='all', inplace=True)
    return df.reset_index(drop=True)

def norm_team(v):
    if pd.isna(v): return ''
    return re.sub(r'^\d+\.', '', str(v)).strip()

def classify_prod(v):
    if pd.isna(v): return None
    s = str(v).upper()
    if any(k in s for k in ['TR', 'RT', '변압기']): return '변압기'
    if any(k in s for k in ['HH', 'HG', 'GIS', 'EA', 'MF', 'GT', '차단기']): return '차단기'
    return None

def to_num(v):
    s = str(v).strip().replace(',', '').replace(' ', '')
    if s in ('', '-', 'nan', '#VALUE!', '#REF!'): return 0.0
    try: return float(s)
    except: return 0.0

def nstr(v):
    if pd.isna(v): return ''
    return re.sub(r'\s+', ' ', str(v).strip())

def gh(team):
    return '해외' if '해외' in str(team) else '국내'

# ============================================================
# STEP 1. 데이터 로드
# ============================================================
# 수주 계획
po = load_csv('사업계획_수주_List.csv')
po = po[po['수주월'].isin(Q1_PLAN)].copy()
po['팀명']    = po['팀명'].apply(norm_team)
po['제품군']  = po['제품군'].apply(classify_prod)
po = po[po['제품군'].notna()].copy()
po['금액']    = po['수주금액(억)'].apply(to_num)
po['월']      = po['수주월'].str.extract(r'(\d+)').astype(int)
po['프로젝트'] = po['프로젝트명'].apply(nstr)
po['PJT']     = po['PJT No.'].apply(nstr) if 'PJT No.' in po.columns else ''
po['발주처']   = po['발주처(EPC등)'].apply(nstr)
po['원청']    = po['원청'].apply(nstr)
po['국내해외'] = po['팀명'].apply(gh)

# 수주 실적
ao = load_csv('실적_수주예상.csv')
ao = ao[ao['확정월'].isin(Q1_ACTUAL)].copy()
ao = ao[ao['수주IN'].astype(str).str.strip() == 'IN'].copy()
ao['팀명']    = ao['팀명'].apply(norm_team)
ao['제품군']  = ao['제품군0'].apply(classify_prod)
ao = ao[ao['제품군'].notna()].copy()
ao['금액']    = ao['예상금액(억원)'].apply(to_num)
ao['월']      = ao['확정월'].str.extract(r'년(\d+)월').astype(int)
ao['프로젝트'] = ao['PJT명'].apply(nstr)
ao['PJT']     = ao['PJT번호'].apply(nstr)
ao['발주처']   = ao['발주처'].apply(nstr)
ao['원청']    = ao['원청'].apply(nstr)
ao['국내해외'] = ao['팀명'].apply(gh)

# 매출 계획
pm = load_csv('사업계획_매출List.csv')
pm = pm[pm['월'].isin(Q1_PLAN)].copy()
pm['팀명']    = pm['팀명'].apply(norm_team)
pm['제품군']  = pm['제품군'].apply(classify_prod)
pm = pm[pm['제품군'].notna()].copy()
pm['금액']    = pm['매출금액(억)'].apply(to_num)
pm['월']      = pm['월'].str.extract(r'(\d+)').astype(int)
pm['프로젝트'] = pm['프로젝트명'].apply(nstr)
pm['PJT']     = pm['PJT No.'].apply(nstr)
pm['발주처']   = pm['발주처(EPC등)'].apply(nstr)
pm['원청']    = pm['원청'].apply(nstr)
pm['국내해외'] = pm['팀명'].apply(gh)

# 매출 실적
am = load_csv('실적_매출예상.csv')
am = am[am['매출월'].isin(Q1_ACTUAL)].copy()
am['팀명']    = am['팀명'].apply(norm_team)
am['제품군']  = am['제품군1'].apply(classify_prod)
am = am[am['제품군'].notna()].copy()
am['금액']    = am['매출(억원)'].apply(to_num)
am['월']      = am['매출월'].str.extract(r'년(\d+)월').astype(int)
am['프로젝트'] = am['PJT명'].apply(nstr)
am['PJT']     = am['PJT번호'].apply(nstr)
am['발주처']   = am['발주처'].apply(nstr)
am['원청']    = ''
am['국내해외'] = am['팀명'].apply(gh)

# ============================================================
# STEP 2. 프로젝트 집계
# ============================================================
def agg_proj(df):
    grp = df.groupby(
        ['프로젝트', 'PJT', '발주처', '원청', '제품군', '팀명', '국내해외'],
        dropna=False
    ).agg(금액=('금액', 'sum'), 월목록=('월', lambda x: sorted(x.unique().tolist()))
    ).reset_index()
    grp['금액'] = grp['금액'].round(2)
    return grp

PA = agg_proj(po)  # 수주 계획
AA = agg_proj(ao)  # 수주 실적
PM = agg_proj(pm)  # 매출 계획
AM = agg_proj(am)  # 매출 실적

# ============================================================
# STEP 3. 매칭 + 원인 분류
# ============================================================
def is_v(s): return bool(s and str(s).strip() not in ('', 'nan'))

def match_and_classify(plan_df, actual_df, label):
    pa = plan_df.copy().reset_index(drop=True)
    ac = actual_df.copy().reset_index(drop=True)
    pa['_pid'] = range(len(pa))
    ac['_aid'] = range(len(ac))

    mp, ma = set(), set()   # matched indices
    results = []

    def record(pr, ar, match_type, score=100):
        diff   = round(ar['금액'] - pr['금액'], 2)
        diff_r = round(diff / pr['금액'] * 100, 1) if pr['금액'] != 0 else np.nan
        변동유형 = '금액동일'
        if abs(diff) > 0.05:
            if   diff_r > 30:  변동유형 = '금액증가(30%↑)'
            elif diff_r < -30: 변동유형 = '금액감소(30%↓)'
            else:              변동유형 = f'금액변동({diff_r:+.0f}%)'
        results.append({
            '구분': label, '원인분류': '금액변동' if abs(diff) > 0.05 else '금액동일',
            '매칭유형': match_type, '유사도': score,
            '제품군': pr['제품군'], '국내해외': pr['국내해외'], '팀명': pr['팀명'],
            '프로젝트_계획': pr['프로젝트'], 'PJT_계획': pr['PJT'], '발주처_계획': pr['발주처'],
            '프로젝트_실적': ar['프로젝트'], 'PJT_실적': ar['PJT'],
            '계획금액': pr['금액'], '실적금액': ar['금액'],
            '차이(억)': diff, '차이율(%)': diff_r, '변동유형': 변동유형,
            '비고': ''
        })
        mp.add(pr['_pid']); ma.add(ar['_aid'])

    # ── 1단계: 제품군 내 정확 매칭
    for prod in ['변압기', '차단기']:
        p_sub = pa[(pa['제품군'] == prod) & ~pa['_pid'].isin(mp)]
        a_sub = ac[(ac['제품군'] == prod) & ~ac['_aid'].isin(ma)]

        # 프로젝트명 정확
        for _, pr in p_sub.iterrows():
            if not is_v(pr['프로젝트']): continue
            cand = a_sub[(~a_sub['_aid'].isin(ma)) & (a_sub['프로젝트'] == pr['프로젝트'])]
            if not cand.empty:
                record(pr, cand.iloc[0], '정확_프로젝트명')

        # PJT번호 정확
        p_sub = pa[(pa['제품군'] == prod) & ~pa['_pid'].isin(mp)]
        a_sub = ac[(ac['제품군'] == prod) & ~ac['_aid'].isin(ma)]
        for _, pr in p_sub.iterrows():
            if not is_v(pr['PJT']): continue
            cand = a_sub[(~a_sub['_aid'].isin(ma)) & (a_sub['PJT'] == pr['PJT'])]
            if not cand.empty:
                record(pr, cand.iloc[0], '정확_PJT번호')

        # 발주처 정확 (같은 제품군+국내해외)
        p_sub = pa[(pa['제품군'] == prod) & ~pa['_pid'].isin(mp)]
        a_sub = ac[(ac['제품군'] == prod) & ~ac['_aid'].isin(ma)]
        for _, pr in p_sub.iterrows():
            if not is_v(pr['발주처']): continue
            cand = a_sub[(~a_sub['_aid'].isin(ma)) &
                         (a_sub['발주처'] == pr['발주처']) &
                         (a_sub['국내해외'] == pr['국내해외'])]
            if not cand.empty:
                record(pr, cand.iloc[0], '정확_발주처')

    # ── 2단계: 퍼지 매칭 (제품군 내, 80% 이상 → 명칭변경 추정)
    for prod in ['변압기', '차단기']:
        p_sub = pa[(pa['제품군'] == prod) & ~pa['_pid'].isin(mp)]
        a_sub = ac[(ac['제품군'] == prod) & ~ac['_aid'].isin(ma)]
        a_names = a_sub['프로젝트'].tolist()
        a_ids   = a_sub['_aid'].tolist()

        for _, pr in p_sub.iterrows():
            if not is_v(pr['프로젝트']): continue
            best_sc, best_i = 0, None
            for i, (an, aid) in enumerate(zip(a_names, a_ids)):
                if aid in ma or not is_v(an): continue
                sc = fuzz.token_sort_ratio(pr['프로젝트'], an)
                if sc > best_sc: best_sc, best_i = sc, i
            if best_sc >= 80 and best_i is not None:
                ar = a_sub[a_sub['_aid'] == a_ids[best_i]].iloc[0]
                r = record.__wrapped__ if hasattr(record, '__wrapped__') else None
                diff   = round(ar['금액'] - pr['금액'], 2)
                diff_r = round(diff / pr['금액'] * 100, 1) if pr['금액'] != 0 else np.nan
                변동유형 = '금액변동' if abs(diff) > 0.05 else '금액동일'
                results.append({
                    '구분': label, '원인분류': '명칭변경추정',
                    '매칭유형': f'유사_{best_sc}%', '유사도': best_sc,
                    '제품군': pr['제품군'], '국내해외': pr['국내해외'], '팀명': pr['팀명'],
                    '프로젝트_계획': pr['프로젝트'], 'PJT_계획': pr['PJT'], '발주처_계획': pr['발주처'],
                    '프로젝트_실적': ar['프로젝트'], 'PJT_실적': ar['PJT'],
                    '계획금액': pr['금액'], '실적금액': ar['금액'],
                    '차이(억)': diff, '차이율(%)': diff_r, '변동유형': 변동유형,
                    '비고': f'유사도 {best_sc}% — 프로젝트명 변경 추정'
                })
                mp.add(pr['_pid']); ma.add(ar['_aid'])

    # ── 3단계: 제품군 교차 매칭 (변압기↔차단기 같은 프로젝트)
    p_rem = pa[~pa['_pid'].isin(mp)]
    a_rem = ac[~ac['_aid'].isin(ma)]
    for _, pr in p_rem.iterrows():
        if not is_v(pr['프로젝트']): continue
        other_prod = '차단기' if pr['제품군'] == '변압기' else '변압기'
        cand = a_rem[(~a_rem['_aid'].isin(ma)) &
                     (a_rem['제품군'] == other_prod) &
                     (a_rem['프로젝트'] == pr['프로젝트'])]
        if not cand.empty:
            ar = cand.iloc[0]
            diff = round(ar['금액'] - pr['금액'], 2)
            diff_r = round(diff / pr['금액'] * 100, 1) if pr['금액'] != 0 else np.nan
            results.append({
                '구분': label, '원인분류': '제품군변경추정',
                '매칭유형': '제품군교차_정확', '유사도': 100,
                '제품군': pr['제품군'], '국내해외': pr['국내해외'], '팀명': pr['팀명'],
                '프로젝트_계획': pr['프로젝트'], 'PJT_계획': pr['PJT'], '발주처_계획': pr['발주처'],
                '프로젝트_실적': ar['프로젝트'], 'PJT_실적': ar['PJT'],
                '계획금액': pr['금액'], '실적금액': ar['금액'],
                '차이(억)': diff, '차이율(%)': diff_r, '변동유형': '제품군변경',
                '비고': f"계획:{pr['제품군']} → 실적:{ar['제품군']}"
            })
            mp.add(pr['_pid']); ma.add(ar['_aid'])

    # ── 4단계: 분할 추정 (계획 1건 → 실적 여러 건)
    p_rem = pa[~pa['_pid'].isin(mp)].copy()
    a_rem = ac[~ac['_aid'].isin(ma)].copy()
    split_used_a = set()

    for _, pr in p_rem.iterrows():
        if not is_v(pr['프로젝트']) or pr['금액'] < 1: continue
        # 실적에서 발주처 동일 or 퍼지 60% 이상인 것 모두 수집
        cands = []
        for _, ar in a_rem[~a_rem['_aid'].isin(split_used_a)].iterrows():
            sc = fuzz.token_sort_ratio(pr['프로젝트'], ar['프로젝트']) if is_v(ar['프로젝트']) else 0
            same_client = (is_v(pr['발주처']) and pr['발주처'] == ar['발주처'])
            if sc >= 60 or same_client:
                cands.append((ar, sc))
        if len(cands) >= 2:
            cand_total = sum(r['금액'] for r, _ in cands)
            diff_pct = abs(cand_total - pr['금액']) / pr['금액'] * 100 if pr['금액'] else 999
            실적목록 = ' / '.join([r['프로젝트'][:20] for r, _ in cands[:5]])
            results.append({
                '구분': label, '원인분류': '분할추정',
                '매칭유형': f'분할_{len(cands)}건', '유사도': max(s for _, s in cands),
                '제품군': pr['제품군'], '국내해외': pr['국내해외'], '팀명': pr['팀명'],
                '프로젝트_계획': pr['프로젝트'], 'PJT_계획': pr['PJT'], '발주처_계획': pr['발주처'],
                '프로젝트_실적': 실적목록,
                'PJT_실적': '',
                '계획금액': pr['금액'],
                '실적금액': round(cand_total, 2),
                '차이(억)': round(cand_total - pr['금액'], 2),
                '차이율(%)': round((cand_total - pr['금액']) / pr['금액'] * 100, 1) if pr['금액'] else np.nan,
                '변동유형': f'분할(실적합계차이{diff_pct:.0f}%)',
                '비고': f"계획 1건 → 실적 {len(cands)}건으로 분할 추정"
            })
            mp.add(pr['_pid'])
            for ar, _ in cands: split_used_a.add(ar['_aid'])

    # ── 5단계: 소멸 (계획에만)
    for _, pr in pa[~pa['_pid'].isin(mp)].iterrows():
        원인 = '소멸_대형(취소/이월추정)' if pr['금액'] >= 10 else '소멸_소형(취소/명칭변경추정)'
        results.append({
            '구분': label, '원인분류': 원인,
            '매칭유형': '소멸', '유사도': 0,
            '제품군': pr['제품군'], '국내해외': pr['국내해외'], '팀명': pr['팀명'],
            '프로젝트_계획': pr['프로젝트'], 'PJT_계획': pr['PJT'], '발주처_계획': pr['발주처'],
            '프로젝트_실적': '', 'PJT_실적': '',
            '계획금액': pr['금액'], '실적금액': 0,
            '차이(억)': -pr['금액'], '차이율(%)': -100.0,
            '변동유형': '소멸', '비고': ''
        })

    # ── 6단계: 신규 (실적에만)
    for _, ar in ac[~ac['_aid'].isin(ma | split_used_a)].iterrows():
        results.append({
            '구분': label, '원인분류': '신규',
            '매칭유형': '신규', '유사도': 0,
            '제품군': ar['제품군'], '국내해외': ar['국내해외'], '팀명': ar['팀명'],
            '프로젝트_계획': '', 'PJT_계획': '', '발주처_계획': '',
            '프로젝트_실적': ar['프로젝트'], 'PJT_실적': ar['PJT'],
            '계획금액': 0, '실적금액': ar['금액'],
            '차이(억)': ar['금액'], '차이율(%)': np.nan,
            '변동유형': '신규', '비고': ''
        })

    return pd.DataFrame(results)

print('매칭 분석 중...')
df_ord = match_and_classify(PA, AA, '수주')
df_sal = match_and_classify(PM, AM, '매출')
df_all = pd.concat([df_ord, df_sal], ignore_index=True)

# ============================================================
# STEP 4. 출력 요약
# ============================================================
def print_summary(df, label):
    total_plan   = df['계획금액'].sum()
    total_actual = df['실적금액'].sum()
    total_diff   = total_actual - total_plan

    print(f'\n{"="*55}')
    print(f'[{label} 차이 원인 분석]')
    print(f'{"="*55}')
    print(f'  계획 합계: {total_plan:,.1f}억')
    print(f'  실적 합계: {total_actual:,.1f}억')
    print(f'  차이 합계: {total_diff:+,.1f}억')
    print()

    grp = df.groupby('원인분류').agg(
        건수=('차이(억)', 'count'),
        차이합계=('차이(억)', 'sum')
    ).reset_index().sort_values('차이합계', key=abs, ascending=False)
    grp['차이합계'] = grp['차이합계'].round(1)
    print('  [원인별 집계]')
    for _, r in grp.iterrows():
        bar = '▲' if r['차이합계'] >= 0 else '▼'
        print(f"  {bar} {r['원인분류']:25s} {r['건수']:3d}건  {r['차이합계']:+8.1f}억")

    for prod in ['변압기', '차단기']:
        sub = df[df['제품군'] == prod]
        print(f'\n  [{prod}]  계획 {sub["계획금액"].sum():.1f}억  → 실적 {sub["실적금액"].sum():.1f}억  ({sub["차이(억)"].sum():+.1f}억)')
        sub2 = sub.groupby('원인분류')['차이(억)'].sum().round(1).sort_values(key=abs, ascending=False)
        for cat, val in sub2.items():
            print(f'    • {cat}: {val:+.1f}억')

    print(f'\n  [차이 상위 10개 프로젝트]')
    top = df.reindex(df['차이(억)'].abs().nlargest(10).index)
    for _, r in top.iterrows():
        proj = r['프로젝트_계획'] or r['프로젝트_실적']
        print(f"  {r['제품군'][0]}/{r['국내해외'][0]}  {proj[:30]:30s}  {r['원인분류']:18s}  {r['차이(억)']:+.1f}억")

print_summary(df_ord, '수주')
print_summary(df_sal, '매출')

# ============================================================
# STEP 5. Excel 저장
# ============================================================
# 스타일
H_FILL  = {
    '수주': PatternFill('solid', fgColor='1F4E79'),
    '매출': PatternFill('solid', fgColor='375623'),
}
CAT_FILL = {
    '신규':               PatternFill('solid', fgColor='E2EFDA'),
    '금액동일':            PatternFill('solid', fgColor='FFFFFF'),
    '금액변동':            PatternFill('solid', fgColor='FFF2CC'),
    '명칭변경추정':         PatternFill('solid', fgColor='FCE4D6'),
    '분할추정':             PatternFill('solid', fgColor='DDEBF7'),
    '제품군변경추정':        PatternFill('solid', fgColor='EAD1DC'),
    '소멸_대형(취소/이월추정)': PatternFill('solid', fgColor='FFDDE1'),
    '소멸_소형(취소/명칭변경추정)': PatternFill('solid', fgColor='FFE9E9'),
}
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

def style_cell(cell, fill=None, bold=False, color='000000', size=9, align='center'):
    if fill: cell.fill = fill
    cell.font = Font(bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = THIN

def write_detail_sheet(wb, df, sheet_name, label):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16

    cols = ['원인분류', '매칭유형', '제품군', '국내해외', '팀명',
            '프로젝트_계획', 'PJT_계획', '발주처_계획',
            '프로젝트_실적', 'PJT_실적',
            '계획금액', '실적금액', '차이(억)', '차이율(%)', '변동유형', '비고']
    col_w = [18, 14, 8, 8, 8, 30, 12, 20, 30, 12, 10, 10, 10, 10, 16, 35]

    # 타이틀
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title = ws.cell(1, 1, f'{label} 차이 원인 분석 (1분기)')
    title.font = Font(bold=True, color='FFFFFF', size=12)
    title.fill = H_FILL.get(label, PatternFill('solid', fgColor='2E75B6'))
    title.alignment = Alignment(horizontal='center', vertical='center')

    # 헤더
    for ci, (h, w) in enumerate(zip(cols, col_w), 1):
        cell = ws.cell(2, ci, h)
        style_cell(cell, PatternFill('solid', fgColor='2E75B6'),
                   bold=True, color='FFFFFF', size=9)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 원인분류 순서 정의
    order = ['금액동일', '금액변동', '명칭변경추정', '분할추정', '제품군변경추정',
             '신규', '소멸_대형(취소/이월추정)', '소멸_소형(취소/명칭변경추정)']
    df_sorted = df.copy()
    df_sorted['_ord'] = df_sorted['원인분류'].apply(
        lambda x: order.index(x) if x in order else 99)
    df_sorted = df_sorted.sort_values(['_ord', '제품군', '국내해외', '차이(억)'],
                                       key=lambda c: c.abs() if c.name == '차이(억)' else c,
                                       ascending=[True, True, True, False])

    prev_cat = None
    for ri, (_, row) in enumerate(df_sorted.iterrows(), 3):
        ws.row_dimensions[ri].height = 14
        # 원인분류가 바뀌면 구분선 색 배경
        cat = row['원인분류']
        fill = CAT_FILL.get(cat, PatternFill('solid', fgColor='FFFFFF'))
        for ci, col in enumerate(cols, 1):
            val = row.get(col, '')
            if isinstance(val, float) and np.isnan(val): val = ''
            cell = ws.cell(ri, ci, val if val != 0 or col in ('계획금액','실적금액','차이(억)') else '')
            use_fill = fill
            style_cell(cell, use_fill, size=9,
                       align='left' if col in ('프로젝트_계획','프로젝트_실적','비고','발주처_계획') else 'center')
            # 차이 강조
            if col == '차이(억)' and isinstance(val, (int, float)) and val != '':
                try:
                    if float(val) > 0:  cell.font = Font(bold=True, color='0070C0', size=9)
                    elif float(val) < 0: cell.font = Font(bold=True, color='C00000', size=9)
                except: pass

    return ws

def write_summary_sheet(wb, df_ord, df_sal):
    ws = wb.create_sheet('요약_워터폴', 0)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8

    def write_waterfall(ws, df, start_row, label):
        # 타이틀
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=7)
        tc = ws.cell(start_row, 1, f'[{label} 차이 원인 워터폴]')
        tc.font = Font(bold=True, color='FFFFFF', size=11)
        tc.fill = H_FILL.get(label, PatternFill('solid', fgColor='2E75B6'))
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[start_row].height = 22

        # 헤더
        headers = ['원인', '건수', '금액 합계(억)', '', '변압기(억)', '차단기(억)', '']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(start_row + 1, ci, h)
            style_cell(c, PatternFill('solid', fgColor='D9E1F2'), bold=True, size=9)
        ws.row_dimensions[start_row + 1].height = 14

        order = ['금액동일', '금액변동', '명칭변경추정', '분할추정', '제품군변경추정',
                 '신규', '소멸_대형(취소/이월추정)', '소멸_소형(취소/명칭변경추정)']
        grp = df.groupby('원인분류').agg(
            건수=('차이(억)', 'count'),
            합계=('차이(억)', 'sum'),
            변압기=('차이(억)', lambda x: x[df.loc[x.index, '제품군'] == '변압기'].sum()),
            차단기=('차이(억)', lambda x: x[df.loc[x.index, '제품군'] == '차단기'].sum()),
        ).reindex([o for o in order if o in df['원인분류'].unique()])

        r = start_row + 2
        total_diff = 0
        for cat, row in grp.iterrows():
            ws.row_dimensions[r].height = 14
            fill = CAT_FILL.get(cat, PatternFill('solid', fgColor='FFFFFF'))
            ws.cell(r, 1, cat).font = Font(size=9)
            ws.cell(r, 2, int(row['건수'])).alignment = Alignment(horizontal='center')
            ws.cell(r, 3, round(row['합계'], 1))
            ws.cell(r, 5, round(row['변압기'], 1))
            ws.cell(r, 6, round(row['차단기'], 1))
            for ci in range(1, 8):
                c = ws.cell(r, ci)
                c.fill = fill
                c.border = THIN
                c.font = Font(size=9)
                if ci in (3, 5, 6):
                    c.alignment = Alignment(horizontal='right')
                    val = c.value
                    if val and isinstance(val, float):
                        if val > 0:   c.font = Font(bold=True, color='0070C0', size=9)
                        elif val < 0: c.font = Font(bold=True, color='C00000', size=9)
            total_diff += row['합계']
            r += 1

        # 합계행
        ws.row_dimensions[r].height = 16
        ws.cell(r, 1, '합계').font = Font(bold=True, size=9)
        ws.cell(r, 3, round(total_diff, 1)).font = Font(bold=True, size=9)
        ws.cell(r, 5, round(df[df['제품군']=='변압기']['차이(억)'].sum(), 1)).font = Font(bold=True, size=9)
        ws.cell(r, 6, round(df[df['제품군']=='차단기']['차이(억)'].sum(), 1)).font = Font(bold=True, size=9)
        for ci in range(1, 8):
            ws.cell(r, ci).fill = PatternFill('solid', fgColor='D9E1F2')
            ws.cell(r, ci).border = THIN
        return r + 2

    next_row = write_waterfall(ws, df_ord, 1, '수주')
    write_waterfall(ws, df_sal, next_row + 1, '매출')
    return ws

wb = openpyxl.Workbook()
wb.remove(wb.active)

write_summary_sheet(wb, df_ord, df_sal)
write_detail_sheet(wb, df_ord, '수주_차이원인', '수주')
write_detail_sheet(wb, df_sal, '매출_차이원인', '매출')

# 분할추정만 별도 시트
split = df_all[df_all['원인분류'] == '분할추정']
if not split.empty:
    write_detail_sheet(wb, split, '분할_명칭변경', '분할/명칭변경')

wb.save(OUT)
print(f'\n저장 완료: {OUT}')
