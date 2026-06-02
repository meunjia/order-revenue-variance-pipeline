#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""중전기 월별 수주/매출 차이분석 (1분기)"""

import os, re, sys, io, warnings
import pandas as pd
import numpy as np
import chardet
from rapidfuzz import fuzz
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

matplotlib.rcParams['font.family'] = 'Malgun Gothic'
matplotlib.rcParams['axes.unicode_minus'] = False

# ── 경로
WORK_DIR  = r"C:\Users\user\claudecode\월별차이"
CSV_DIR   = os.path.join(WORK_DIR, "csv")
OUT_XLSX  = os.path.join(WORK_DIR, "차이분석_결과.xlsx")
CHART_DIR = os.path.join(WORK_DIR, "charts")
os.makedirs(CHART_DIR, exist_ok=True)

# ── 실행 모드 ──────────────────────────────────────────────────
# True  = csv/sample_*.csv 로 실행  (GitHub 샘플, 기본값)
# False = 실제 데이터 파일로 실행
SAMPLE_MODE = True

if SAMPLE_MODE:
    FILE_PO = 'sample_사업계획_수주.csv'
    FILE_AO = 'sample_실적_수주.csv'
    FILE_PM = 'sample_사업계획_매출.csv'
    FILE_AM = 'sample_실적_매출.csv'
else:
    FILE_PO = '사업계획_수주_List_123월.csv'
    FILE_AO = '실적_수주예상.csv'
    FILE_PM = '사업계획_매출List.csv'
    FILE_AM = '실적_매출예상.csv'
# ──────────────────────────────────────────────────────────────

Q1_PLAN   = ['01월', '02월', '03월']
Q1_ACTUAL = ['26년01월', '26년02월', '26년03월']
MONTH_LABEL = {1:'1월', 2:'2월', 3:'3월'}

# ============================================================
# 유틸
# ============================================================
def detect_enc(path):
    with open(path, 'rb') as f:
        return chardet.detect(f.read(100000))['encoding'] or 'utf-8-sig'

def load_csv(fname):
    path = os.path.join(CSV_DIR, fname)
    enc = detect_enc(path)
    df = pd.read_csv(path, encoding=enc)
    df.columns = [str(c).strip().replace('\n', '') for c in df.columns]
    df.dropna(how='all', inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df

def norm_team(v):
    if pd.isna(v): return None
    return re.sub(r'^\d+\.', '', str(v)).strip()

def classify_prod(v):
    if pd.isna(v): return None
    s = str(v).upper().strip()
    if any(k in s for k in ['TR', 'RT', '변압기']): return '변압기'
    if any(k in s for k in ['HH', 'HG', 'GIS', 'EA', 'MF', 'GT', '차단기']): return '차단기'
    return None

def to_num(v):
    if pd.isna(v): return 0.0
    s = str(v).strip().replace(',', '').replace(' ', '')
    if s in ('', '-', 'nan', '#VALUE!', '#REF!'): return 0.0
    try: return float(s)
    except: return 0.0

def nstr(v):
    if pd.isna(v): return ''
    return re.sub(r'\s+', ' ', str(v).strip())

def is_valid(s):
    return bool(s and s.strip() and s not in ('nan', 'NaN'))

def team_to_gh(t):
    if t is None: return '국내'
    return '해외' if '해외' in str(t) else '국내'

# ============================================================
# STEP 1. 데이터 로드
# ============================================================
print('\n' + '='*60)
print('STEP 1. 데이터 로드 및 전처리')
print('='*60)

# ── 1-1. 수주 계획 (상단 9줄 = 환율/구성비 헤더 → skip)
_path = os.path.join(CSV_DIR, FILE_PO)
_enc  = detect_enc(_path)
raw = pd.read_csv(_path, encoding=_enc, skiprows=list(range(9)), header=0, low_memory=False)
raw.columns = [str(c).strip().replace('\n', '') for c in raw.columns]
raw.dropna(how='all', inplace=True)
raw.reset_index(drop=True, inplace=True)
print(f'\n[{FILE_PO}] 원본 {len(raw):,}행  (skiprows=9 적용)')
df_po = raw[raw['수주월'].isin(Q1_PLAN)].copy()
df_po['팀명']   = df_po['팀명'].apply(norm_team)
df_po['제품군'] = df_po['제품군'].apply(classify_prod)
df_po = df_po[df_po['제품군'].notna()].copy()
df_po['금액']   = df_po['수주금액(억)'].apply(to_num)
df_po['월']     = df_po['수주월'].str.extract(r'(\d+)').astype(int)
df_po['프로젝트'] = df_po['프로젝트명'].apply(nstr)
df_po['PJT']    = df_po['PJT No.'].apply(nstr) if 'PJT No.' in df_po.columns else ''
df_po['발주처']  = df_po['발주처(EPC등)'].apply(nstr)
df_po['원청']   = df_po['원청'].apply(nstr)
df_po['국내해외'] = df_po['팀명'].apply(team_to_gh)
print(f'  → 1분기 필터 후 {len(df_po):,}행')
print(df_po[['월','팀명','제품군','금액','프로젝트']].head(5).to_string(index=False))

# ── 1-2. 수주 실적
raw = load_csv(FILE_AO)
print(f'\n[{FILE_AO}] 원본 {len(raw):,}행')
df_ao = raw[raw['확정월'].isin(Q1_ACTUAL)].copy()
df_ao = df_ao[df_ao['수주IN'].astype(str).str.strip() == 'IN'].copy()
df_ao['팀명']   = df_ao['팀명'].apply(norm_team)
df_ao['제품군'] = df_ao['제품군0'].apply(classify_prod)
df_ao = df_ao[df_ao['제품군'].notna()].copy()
df_ao['금액']   = df_ao['예상금액(억원)'].apply(to_num)
df_ao['월']     = df_ao['확정월'].str.extract(r'년(\d+)월').astype(int)
df_ao['프로젝트'] = df_ao['PJT명'].apply(nstr)
df_ao['PJT']    = df_ao['PJT번호'].apply(nstr)
df_ao['발주처']  = df_ao['발주처'].apply(nstr)
df_ao['원청']   = df_ao['원청'].apply(nstr)
df_ao['국내해외'] = df_ao['팀명'].apply(team_to_gh)
print(f'  → 1분기+IN 필터 후 {len(df_ao):,}행')
print(df_ao[['월','팀명','제품군','금액','프로젝트']].head(5).to_string(index=False))

# ── 1-3. 매출 계획
raw = load_csv(FILE_PM)
print(f'\n[{FILE_PM}] 원본 {len(raw):,}행')
df_pm = raw[raw['월'].isin(Q1_PLAN)].copy()
df_pm['팀명']   = df_pm['팀명'].apply(norm_team)
df_pm['제품군'] = df_pm['제품군'].apply(classify_prod)
df_pm = df_pm[df_pm['제품군'].notna()].copy()
df_pm['금액']   = df_pm['매출금액(억)'].apply(to_num)
df_pm['월']     = df_pm['월'].str.extract(r'(\d+)').astype(int)
df_pm['프로젝트'] = df_pm['프로젝트명'].apply(nstr)
df_pm['PJT']    = df_pm['PJT No.'].apply(nstr)
df_pm['발주처']  = df_pm['발주처(EPC등)'].apply(nstr)
df_pm['원청']   = df_pm['원청'].apply(nstr)
df_pm['국내해외'] = df_pm['팀명'].apply(team_to_gh)
print(f'  → 1분기 필터 후 {len(df_pm):,}행')
print(df_pm[['월','팀명','제품군','금액','프로젝트']].head(5).to_string(index=False))

# ── 1-4. 매출 실적 (상단 4줄 = 환율 헤더 → skip, 5행부터 컬럼명)
_path4 = os.path.join(CSV_DIR, FILE_AM)
_enc4  = detect_enc(_path4)
raw = pd.read_csv(_path4, encoding=_enc4, skiprows=list(range(4)), header=0, low_memory=False)
raw.columns = [str(c).strip().replace('\n', '') for c in raw.columns]
raw.dropna(how='all', inplace=True)
raw.reset_index(drop=True, inplace=True)
print(f'\n[{FILE_AM}] 원본 {len(raw):,}행  (skiprows=4 적용)')
df_am = raw[raw['매출월'].isin(Q1_ACTUAL)].copy()
df_am['팀명']   = df_am['팀명'].apply(norm_team)
df_am['제품군'] = df_am['제품군1'].apply(classify_prod)
df_am = df_am[df_am['제품군'].notna()].copy()
df_am['금액']   = df_am['매출(억원)'].apply(to_num)
df_am['월']     = df_am['매출월'].str.extract(r'년(\d+)월').astype(int)
df_am['프로젝트'] = df_am['PJT명'].apply(nstr)
df_am['PJT']    = df_am['PJT번호'].apply(nstr)
df_am['발주처']  = df_am['발주처'].apply(nstr)
df_am['원청']   = ''
df_am['국내해외'] = df_am['팀명'].apply(team_to_gh)
print(f'  → 1분기 필터 후 {len(df_am):,}행')
print(df_am[['월','팀명','제품군','금액','프로젝트']].head(5).to_string(index=False))

# ============================================================
# STEP 2. 요약 비교
# ============================================================
print('\n' + '='*60)
print('STEP 2. 요약 비교 (팀명/제품군/월 groupby)')
print('='*60)

KEY = ['월', '팀명', '제품군']

def summary_compare(df_plan, df_actual, label):
    gp = df_plan.groupby(KEY)['금액'].sum().reset_index().rename(columns={'금액':'계획'})
    ga = df_actual.groupby(KEY)['금액'].sum().reset_index().rename(columns={'금액':'실적'})
    mg = pd.merge(gp, ga, on=KEY, how='outer').fillna(0)
    mg['차이(억)'] = (mg['실적'] - mg['계획']).round(2)
    mg['달성률(%)'] = np.where(mg['계획'] != 0,
                               (mg['실적'] / mg['계획'] * 100).round(1), np.nan)
    mg = mg.sort_values(['제품군','월','팀명']).reset_index(drop=True)
    print(f'\n[{label} 요약비교]')
    print(mg.to_string(index=False))
    print(f'\n  ▶ 차이 상위 5개:')
    top5 = mg.reindex(mg['차이(억)'].abs().nlargest(5).index)
    print(top5[KEY + ['계획','실적','차이(억)','달성률(%)']].to_string(index=False))
    return mg

sum_order  = summary_compare(df_po, df_ao, '수주')
sum_sales  = summary_compare(df_pm, df_am, '매출')

# ── 전체 합계 (Excel 채우기용)
def q1_total(df):
    return round(df['금액'].sum(), 2)

totals = {
    '수주계획_중전기':  q1_total(df_po),
    '수주실적_중전기':  q1_total(df_ao),
    '수주계획_변압기':  q1_total(df_po[df_po['제품군']=='변압기']),
    '수주실적_변압기':  q1_total(df_ao[df_ao['제품군']=='변압기']),
    '수주계획_차단기':  q1_total(df_po[df_po['제품군']=='차단기']),
    '수주실적_차단기':  q1_total(df_ao[df_ao['제품군']=='차단기']),
    '매출계획_중전기':  q1_total(df_pm),
    '매출실적_중전기':  q1_total(df_am),
    '매출계획_변압기':  q1_total(df_pm[df_pm['제품군']=='변압기']),
    '매출실적_변압기':  q1_total(df_am[df_am['제품군']=='변압기']),
    '매출계획_차단기':  q1_total(df_pm[df_pm['제품군']=='차단기']),
    '매출실적_차단기':  q1_total(df_am[df_am['제품군']=='차단기']),
}
print('\n[1분기 합계 (억)]')
for k, v in totals.items(): print(f'  {k}: {v}')

# ============================================================
# STEP 3. 프로젝트 심층 비교
# ============================================================
print('\n' + '='*60)
print('STEP 3. 프로젝트 심층 비교')
print('='*60)

def aggregate_proj(df):
    """프로젝트 단위 집계"""
    grp_cols = ['프로젝트','PJT','발주처','원청','제품군','팀명','국내해외']
    agg = df.groupby(grp_cols, dropna=False)['금액'].sum().reset_index()
    return agg

def match_projects(plan_df, actual_df, category_name):
    """
    매칭 우선순위: 프로젝트명 → PJT번호 → 발주처 → 원청 → 퍼지(80%)
    Returns: matched_df, unmatched_plan_df, unmatched_actual_df
    """
    pa = plan_df.copy().reset_index(drop=True)
    ac = actual_df.copy().reset_index(drop=True)
    pa['_pid'] = range(len(pa))
    ac['_aid'] = range(len(ac))

    matched_p = set()
    matched_a = set()
    results = []

    def do_match(key_col, match_type):
        for _, pr in pa[~pa['_pid'].isin(matched_p)].iterrows():
            kv = pr[key_col]
            if not is_valid(kv): continue
            cands = ac[(~ac['_aid'].isin(matched_a)) &
                       (ac['제품군'] == pr['제품군']) &
                       (ac[key_col] == kv)]
            if cands.empty: continue
            ar = cands.iloc[0]
            diff = round(ar['금액'] - pr['금액'], 2)
            pct  = round(diff / pr['금액'] * 100, 1) if pr['금액'] != 0 else np.nan
            comment = ''
            if not np.isnan(pct) and abs(pct) >= 30:
                comment = '단가 또는 수량 변동 추정'
            results.append({
                '구분': category_name, '매칭유형': match_type, '유사도': 100,
                '제품군': pr['제품군'], '국내해외': pr['국내해외'],
                '팀명': pr['팀명'],
                '프로젝트_계획': pr['프로젝트'], 'PJT_계획': pr['PJT'],
                '발주처_계획': pr['발주처'],
                '프로젝트_실적': ar['프로젝트'],
                '계획금액': round(pr['금액'], 2),
                '실적금액': round(ar['금액'], 2),
                '차이(억)': diff, '차이율(%)': pct, '코멘트': comment,
            })
            matched_p.add(pr['_pid'])
            matched_a.add(ar['_aid'])

    do_match('프로젝트', '정확_프로젝트명')
    do_match('PJT',    '정확_PJT번호')
    do_match('발주처',  '정확_발주처')
    do_match('원청',   '정확_원청')

    # 퍼지 매칭 (프로젝트명, 80% 이상)
    remaining_p = pa[~pa['_pid'].isin(matched_p)].copy()
    remaining_a = ac[~ac['_aid'].isin(matched_a)].copy()
    for prod in ['변압기', '차단기']:
        rp = remaining_p[remaining_p['제품군'] == prod]
        ra = remaining_a[remaining_a['제품군'] == prod]
        if rp.empty or ra.empty: continue
        cand_names = ra['프로젝트'].tolist()
        for _, pr in rp.iterrows():
            if not is_valid(pr['프로젝트']): continue
            best_name, best_score, best_idx = None, 0, None
            for i, cn in enumerate(cand_names):
                if not is_valid(cn): continue
                sc = fuzz.token_sort_ratio(pr['프로젝트'], cn)
                if sc > best_score:
                    best_score, best_name, best_idx = sc, cn, i
            if best_score >= 80 and best_idx is not None:
                ar_row = ra[ra['프로젝트'] == best_name]
                if ar_row.empty: continue
                ar = ar_row.iloc[0]
                if ar['_aid'] in matched_a: continue
                diff = round(ar['금액'] - pr['금액'], 2)
                pct  = round(diff / pr['금액'] * 100, 1) if pr['금액'] != 0 else np.nan
                comment = ''
                if not np.isnan(pct) and abs(pct) >= 30:
                    comment = '단가 또는 수량 변동 추정'
                results.append({
                    '구분': category_name, '매칭유형': '유사_프로젝트명',
                    '유사도': best_score,
                    '제품군': pr['제품군'], '국내해외': pr['국내해외'],
                    '팀명': pr['팀명'],
                    '프로젝트_계획': pr['프로젝트'], 'PJT_계획': pr['PJT'],
                    '발주처_계획': pr['발주처'],
                    '프로젝트_실적': ar['프로젝트'],
                    '계획금액': round(pr['금액'], 2),
                    '실적금액': round(ar['금액'], 2),
                    '차이(억)': diff, '차이율(%)': pct, '코멘트': comment,
                })
                matched_p.add(pr['_pid'])
                matched_a.add(ar['_aid'])

    matched_df = pd.DataFrame(results) if results else pd.DataFrame()

    # 소멸 (계획에만)
    unm_p = pa[~pa['_pid'].isin(matched_p)].copy()
    unm_p['분류'] = unm_p['금액'].apply(
        lambda x: '소멸_대형(취소/이월추정)' if x >= 10 else '소멸_소형(취소/명칭변경추정)')

    # 신규 (실적에만)
    unm_a = ac[~ac['_aid'].isin(matched_a)].copy()
    unm_a['분류'] = unm_a['프로젝트'].apply(
        lambda x: '신규_계획외' if is_valid(x) else '신규_기존분리추정')

    return matched_df, unm_p, unm_a

# 수주
pa_order = aggregate_proj(df_po)
aa_order = aggregate_proj(df_ao)
m_ord, unp_ord, una_ord = match_projects(pa_order, aa_order, '수주')

# 매출
pa_sales = aggregate_proj(df_pm)
aa_sales = aggregate_proj(df_am)
m_sal, unp_sal, una_sal = match_projects(pa_sales, aa_sales, '매출')

def print_step3(label, matched, unm_p, unm_a):
    print(f'\n[{label} 매칭 결과]')
    print(f'  매칭: {len(matched)}건 / 소멸: {len(unm_p)}건 / 신규: {len(unm_a)}건')
    if not matched.empty:
        top5 = matched.reindex(matched['차이(억)'].abs().nlargest(5).index)
        print(f'  ▶ 차이 상위 5:')
        print(top5[['제품군','국내해외','프로젝트_계획','계획금액','실적금액','차이(억)','코멘트']].to_string(index=False))
    if not unm_p.empty:
        top_p = unm_p.nlargest(5, '금액')
        print(f'  ▶ 소멸 상위 5:')
        print(top_p[['제품군','국내해외','프로젝트','금액','분류']].to_string(index=False))
    if not unm_a.empty:
        top_a = unm_a.nlargest(5, '금액')
        print(f'  ▶ 신규 상위 5:')
        print(top_a[['제품군','국내해외','프로젝트','금액','분류']].to_string(index=False))

print_step3('수주', m_ord, unp_ord, una_ord)
print_step3('매출', m_sal, unp_sal, una_sal)

# ============================================================
# STEP 4. 결과물 생성
# ============================================================
print('\n' + '='*60)
print('STEP 4. 결과물 생성')
print('='*60)

# ── 4-1. 신규_소멸 통합 DF
def build_news_extinct(unm_p, unm_a, label):
    rows = []
    for _, r in unm_p.iterrows():
        rows.append({'구분': label, '분류': r['분류'], '제품군': r['제품군'],
                     '국내해외': r['국내해외'], '팀명': r['팀명'],
                     '프로젝트': r['프로젝트'], 'PJT': r['PJT'],
                     '발주처': r['발주처'], '금액': round(r['금액'], 2)})
    for _, r in unm_a.iterrows():
        rows.append({'구분': label, '분류': r['분류'], '제품군': r['제품군'],
                     '국내해외': r['국내해외'], '팀명': r['팀명'],
                     '프로젝트': r['프로젝트'], 'PJT': r['PJT'],
                     '발주처': r['발주처'], '금액': round(r['금액'], 2)})
    return pd.DataFrame(rows)

ne_df = pd.concat([build_news_extinct(unp_ord, una_ord, '수주'),
                   build_news_extinct(unp_sal, una_sal, '매출')], ignore_index=True)

# ── 4-2. 텍스트 사유 생성
def gen_reason_text(matched, unm_p, unm_a, label):
    """Step3 결과로 텍스트 사유 생성"""
    lines = []

    # 소멸 (계획에만, 금액 내림차순)
    for _, r in unm_p.sort_values('금액', ascending=False).iterrows():
        if r['금액'] < 0.1: continue
        prefix = ('변' if r['제품군']=='변압기' else '차')
        gh     = ('국' if r['국내해외']=='국내' else '해')
        sign   = '-'
        lines.append(f"{prefix}/{gh}) {r['프로젝트']} {sign}{abs(r['금액']):.1f}억 (취소/이월 추정)")

    # 신규 (실적에만, 금액 내림차순)
    for _, r in unm_a.sort_values('금액', ascending=False).iterrows():
        if r['금액'] < 0.1: continue
        prefix = ('변' if r['제품군']=='변압기' else '차')
        gh     = ('국' if r['국내해외']=='국내' else '해')
        lines.append(f"{prefix}/{gh}) {r['프로젝트']} +{r['금액']:.1f}억 (신규)")

    # 금액 변동 (차이 절대값 내림차순)
    if not matched.empty:
        for _, r in matched.sort_values('차이(억)', key=abs, ascending=False).iterrows():
            if abs(r['차이(억)']) < 0.1: continue
            prefix = ('변' if r['제품군']=='변압기' else '차')
            gh     = ('국' if r['국내해외']=='국내' else '해')
            sign   = '+' if r['차이(억)'] >= 0 else ''
            lines.append(f"{prefix}/{gh}) {r['프로젝트_계획']} {sign}{r['차이(억)']:.1f}억")

    total_diff = round(
        una_a_sum(unm_a) - unm_p['금액'].sum() +
        (matched['차이(억)'].sum() if not matched.empty else 0), 1)

    print(f'\n[{label} 차이 텍스트 사유]')
    print(f'  합계 차이: {total_diff:+.1f}억')
    for l in lines[:20]:
        print(f'  {l}')
    return total_diff, lines

def una_a_sum(df):
    return df['금액'].sum() if not df.empty else 0.0

ord_diff, ord_lines = gen_reason_text(m_ord, unp_ord, una_ord, '수주')
sal_diff, sal_lines = gen_reason_text(m_sal, unp_sal, una_sal, '매출')

# ── 4-3. 시각화
print('\n[시각화 생성 중...]')

BLUE   = '#2E75B6'
ORANGE = '#ED7D31'
GREEN  = '#70AD47'
RED    = '#FF0000'
GRAY   = '#A6A6A6'

def save_chart(fig, name):
    path = os.path.join(CHART_DIR, name)
    fig.savefig(path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return path

saved_charts = {}

# Chart 1 & 2: 수주/매출 계획 vs 실적 (변압기, 차단기 각각)
def monthly_bar_chart(sum_df, title, filename):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    fig.suptitle(title, fontsize=14, fontweight='bold')
    for ax, prod in zip(axes, ['변압기', '차단기']):
        sub = sum_df[sum_df['제품군'] == prod].copy()
        months = [1, 2, 3]
        plan_vals   = [sub[sub['월']==m]['계획'].sum() for m in months]
        actual_vals = [sub[sub['월']==m]['실적'].sum() for m in months]
        x = np.arange(3)
        w = 0.35
        ax.bar(x - w/2, plan_vals,   w, label='계획', color=BLUE,   alpha=0.85)
        ax.bar(x + w/2, actual_vals, w, label='실적', color=ORANGE, alpha=0.85)
        ax.set_title(prod, fontsize=12)
        ax.set_xticks(x)
        ax.set_xticklabels(['1월','2월','3월'])
        ax.set_ylabel('금액 (억)')
        ax.legend()
        ax.grid(axis='y', linestyle='--', alpha=0.4)
        for i, (p, a) in enumerate(zip(plan_vals, actual_vals)):
            ax.text(i-w/2, p+0.5, f'{p:.0f}', ha='center', va='bottom', fontsize=8)
            ax.text(i+w/2, a+0.5, f'{a:.0f}', ha='center', va='bottom', fontsize=8)
    plt.tight_layout()
    return save_chart(fig, filename)

saved_charts['수주_월별'] = monthly_bar_chart(sum_order, '수주 계획 vs 실적 (월별)', '수주_월별.png')
saved_charts['매출_월별'] = monthly_bar_chart(sum_sales, '매출 계획 vs 실적 (월별)', '매출_월별.png')

# Chart 3: 수주 워터폴 차트
def waterfall_chart(계획, 신규합, 소멸합, 변동합, 실적, title, filename):
    labels = ['계획', '신규(+)', '소멸(-)', '금액변동', '실적']
    values = [계획, 신규합, -소멸합, 변동합, 실적]
    running = [0, 계획, 계획+신규합, 계획+신규합-소멸합, 0]
    colors = [BLUE, GREEN, RED, (GREEN if 변동합>=0 else RED), BLUE]

    fig, ax = plt.subplots(figsize=(10, 6))
    fig.suptitle(title, fontsize=13, fontweight='bold')
    for i, (lb, val, bot, col) in enumerate(zip(labels, values, running, colors)):
        ax.bar(i, abs(val), bottom=(bot if i not in (0,4) else 0), color=col, alpha=0.85, width=0.5)
        ypos = (bot + val/2) if i not in (0,4) else abs(val)/2
        ax.text(i, (bot+abs(val)+2), f'{val:+.0f}억', ha='center', va='bottom', fontsize=9, fontweight='bold')
    ax.set_xticks(range(5))
    ax.set_xticklabels(labels)
    ax.set_ylabel('금액 (억)')
    ax.axhline(0, color='black', linewidth=0.5)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    plt.tight_layout()
    return save_chart(fig, filename)

신규수주합 = una_ord['금액'].sum()
소멸수주합 = unp_ord['금액'].sum()
변동수주합 = m_ord['차이(억)'].sum() if not m_ord.empty else 0

saved_charts['수주_워터폴'] = waterfall_chart(
    totals['수주계획_중전기'], 신규수주합, 소멸수주합, 변동수주합, totals['수주실적_중전기'],
    '수주 워터폴 (1분기)', '수주_워터폴.png')

# Chart 4: 프로젝트별 차이 Top 10
def top10_project_chart(matched, unm_p, unm_a, title, filename):
    rows = []
    if not matched.empty:
        for _, r in matched.iterrows():
            rows.append({'프로젝트': r['프로젝트_계획'][:20], '차이': r['차이(억)'],
                         '제품군': r['제품군'], '국내해외': r['국내해외']})
    for _, r in unm_a.iterrows():
        rows.append({'프로젝트': r['프로젝트'][:20], '차이': r['금액'],
                     '제품군': r['제품군'], '국내해외': r['국내해외']})
    for _, r in unm_p.iterrows():
        rows.append({'프로젝트': r['프로젝트'][:20], '차이': -r['금액'],
                     '제품군': r['제품군'], '국내해외': r['국내해외']})
    if not rows:
        return None
    df_t = pd.DataFrame(rows).sort_values('차이', key=abs, ascending=False).head(10)
    fig, ax = plt.subplots(figsize=(12, 6))
    fig.suptitle(title, fontsize=13, fontweight='bold')
    labels = [f"[{r['제품군'][0]}/{r['국내해외'][0]}] {r['프로젝트']}" for _, r in df_t.iterrows()]
    vals   = df_t['차이'].tolist()
    colors = [GREEN if v >= 0 else RED for v in vals]
    y = range(len(labels))
    ax.barh(list(y), vals, color=colors, alpha=0.85)
    ax.set_yticks(list(y))
    ax.set_yticklabels(labels, fontsize=9)
    ax.axvline(0, color='black', linewidth=0.8)
    ax.set_xlabel('차이 금액 (억)')
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    for i, v in enumerate(vals):
        ax.text(v + (1 if v>=0 else -1), i, f'{v:+.1f}억', va='center', fontsize=8)
    plt.tight_layout()
    return save_chart(fig, filename)

saved_charts['수주_top10'] = top10_project_chart(m_ord, unp_ord, una_ord,
    '수주 프로젝트별 차이 Top10', '수주_top10.png')
saved_charts['매출_top10'] = top10_project_chart(m_sal, unp_sal, una_sal,
    '매출 프로젝트별 차이 Top10', '매출_top10.png')

# Chart 5: 팀별 달성률 도넛
def team_donut_chart(sum_df, label, filename):
    teams = ['국내1','국내2','해외1','해외2']
    fig, axes = plt.subplots(1, 4, figsize=(16, 5))
    fig.suptitle(f'{label} 팀별 달성률', fontsize=13, fontweight='bold')
    for ax, team in zip(axes, teams):
        sub = sum_df[sum_df['팀명'] == team]
        plan = sub['계획'].sum()
        act  = sub['실적'].sum()
        rate = round(act/plan*100, 1) if plan > 0 else 0
        done  = min(rate, 100)
        left  = max(0, 100 - done)
        over  = max(0, rate - 100)
        if over > 0:
            wedges, _ = ax.pie([100, over], colors=[BLUE, GREEN], startangle=90,
                               wedgeprops=dict(width=0.4))
        else:
            wedges, _ = ax.pie([done, left], colors=[BLUE, GRAY], startangle=90,
                               wedgeprops=dict(width=0.4))
        ax.text(0, 0, f'{rate}%', ha='center', va='center', fontsize=14, fontweight='bold',
                color=(GREEN if rate>=100 else BLUE))
        ax.set_title(f'{team}\n계획 {plan:.0f}억\n실적 {act:.0f}억', fontsize=9)
    plt.tight_layout()
    return save_chart(fig, filename)

saved_charts['수주_팀달성률'] = team_donut_chart(sum_order, '수주', '수주_팀달성률.png')
saved_charts['매출_팀달성률'] = team_donut_chart(sum_sales, '매출', '매출_팀달성률.png')

print('  차트 저장 완료:', list(saved_charts.keys()))

# ── 4-4. 차이분석_결과.xlsx 저장
print('\n[차이분석_결과.xlsx 저장 중...]')

HDR_FILL  = PatternFill('solid', fgColor='2E75B6')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
DIFF_PLUS = PatternFill('solid', fgColor='E2EFDA')
DIFF_MINUS= PatternFill('solid', fgColor='FFDDE1')
BOLD_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def write_df_to_sheet(ws, df, start_row=1):
    """DataFrame을 시트에 쓰기"""
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    for ri, row in enumerate(df.itertuples(index=False), start_row+1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            # 차이 컬럼 색상
            if df.columns[ci-1] in ('차이(억)', '차이율(%)'):
                try:
                    if float(val) > 0:  cell.fill = DIFF_PLUS
                    elif float(val) < 0: cell.fill = DIFF_MINUS
                except: pass
    # 컬럼 너비 자동
    for ci, col in enumerate(df.columns, 1):
        max_len = max([len(str(col))] +
                      [len(str(ws.cell(r, ci).value or '')) for r in range(start_row, ws.max_row+1)])
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# 시트1: 수주_요약비교
ws = wb.create_sheet('수주_요약비교')
write_df_to_sheet(ws, sum_order)

# 시트2: 매출_요약비교
ws = wb.create_sheet('매출_요약비교')
write_df_to_sheet(ws, sum_sales)

# 시트3: 수주_프로젝트심층
ws = wb.create_sheet('수주_프로젝트심층')
if not m_ord.empty:
    write_df_to_sheet(ws, m_ord)

# 시트4: 매출_프로젝트심층
ws = wb.create_sheet('매출_프로젝트심층')
if not m_sal.empty:
    write_df_to_sheet(ws, m_sal)

# 시트5: 신규_소멸_프로젝트
ws = wb.create_sheet('신규_소멸_프로젝트')
if not ne_df.empty:
    write_df_to_sheet(ws, ne_df)

# 시트6: 텍스트_사유
ws = wb.create_sheet('텍스트_사유')
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 80
rows_txt = [
    ('항목', '내용'),
    ('수주 합계 차이', f'{totals["수주실적_중전기"]-totals["수주계획_중전기"]:+.1f}억'),
    ('매출 합계 차이', f'{totals["매출실적_중전기"]-totals["매출계획_중전기"]:+.1f}억'),
    ('', ''),
    ('== 수주 사유 ==', ''),
]
for l in ord_lines: rows_txt.append(('수주', l))
rows_txt.append(('', ''))
rows_txt.append(('== 매출 사유 ==', ''))
for l in sal_lines: rows_txt.append(('매출', l))
for ri, (a, b) in enumerate(rows_txt, 1):
    ws.cell(ri, 1, a); ws.cell(ri, 2, b)
    if ri == 1:
        for ci in (1,2):
            ws.cell(ri,ci).font = HDR_FONT
            ws.cell(ri,ci).fill = HDR_FILL

# 시트7: 시각화_대시보드
ws = wb.create_sheet('시각화_대시보드')
ws.sheet_view.showGridLines = False
titles_map = [
    ('수주_월별',     'B2',  '▶ 수주 계획 vs 실적 (월별)'),
    ('매출_월별',     'B32', '▶ 매출 계획 vs 실적 (월별)'),
    ('수주_워터폴',   'P2',  '▶ 수주 워터폴 차트'),
    ('수주_top10',   'B62', '▶ 수주 Top10 프로젝트 차이'),
    ('매출_top10',   'P62', '▶ 매출 Top10 프로젝트 차이'),
    ('수주_팀달성률', 'B92', '▶ 수주 팀별 달성률'),
    ('매출_팀달성률', 'P92', '▶ 매출 팀별 달성률'),
]
for key, anchor, title in titles_map:
    path = saved_charts.get(key)
    if path and os.path.exists(path):
        row = int(anchor[1:])
        col = anchor[0]
        ws.cell(row, openpyxl.utils.column_index_from_string(col),
                value=title).font = BOLD_FONT
        img = XLImage(path)
        img.anchor = f'{col}{row+1}'
        ws.add_image(img)

wb.save(OUT_XLSX)
print(f'  → 저장: {OUT_XLSX}')

# ── 4-5. 중전기_월별차이분석 Excel 채우기 시도
ORIG_XLSX = os.path.join(CSV_DIR, '중전기_월별차이분석_20260422.xlsx')
print('\n[중전기_월별차이분석_20260422.xlsx 수치 자동 채우기 시도...]')
try:
    wb2 = openpyxl.load_workbook(ORIG_XLSX)
    ws2 = wb2.active

    # 구조: col C=1분기계획, D=1분기실적
    # 행 번호는 merged cell 때문에 헤더 포함 실제 위치 확인 필요
    # CSV 기준 row3~11 → Excel row 4~12 (헤더 3행)
    fill_map = {
        # (excel_row, col_계획, col_실적): (계획값, 실적값)
        (4,  3, 4): (totals['수주계획_중전기'],  totals['수주실적_중전기']),   # 중전기 수주
        (5,  3, 4): (totals['매출계획_중전기'],  totals['매출실적_중전기']),   # 중전기 매출
        (7,  3, 4): (totals['수주계획_변압기'],  totals['수주실적_변압기']),   # 변압기 수주
        (8,  3, 4): (totals['매출계획_변압기'],  totals['매출실적_변압기']),   # 변압기 매출
        (10, 3, 4): (totals['수주계획_차단기'],  totals['수주실적_차단기']),   # 차단기 수주
        (11, 3, 4): (totals['매출계획_차단기'],  totals['매출실적_차단기']),   # 차단기 매출
    }
    for (r, cp, ca), (plan_v, act_v) in fill_map.items():
        ws2.cell(r, cp, round(plan_v, 1))
        ws2.cell(r, ca, round(act_v,  1))
    wb2.save(ORIG_XLSX)
    print('  → 수치 채우기 성공!')
except Exception as e:
    print(f'  → DRM 보호로 직접 수정 불가: {e}')
    print('\n  [수동 입력 값 (억)]')
    print(f'  ┌──────────────┬──────────┬──────────┐')
    print(f'  │ 항목         │   계획   │   실적   │')
    print(f'  ├──────────────┼──────────┼──────────┤')
    for cat in ['중전기', '변압기', '차단기']:
        for item in ['수주', '매출']:
            p = totals[f'{item}계획_{cat}']
            a = totals[f'{item}실적_{cat}']
            print(f'  │ {cat} {item}  │ {p:8.1f} │ {a:8.1f} │')
    print(f'  └──────────────┴──────────┴──────────┘')

print('\n' + '='*60)
print('완료!')
print(f'  결과 파일: {OUT_XLSX}')
print(f'  차트 폴더: {CHART_DIR}')
print('='*60)
